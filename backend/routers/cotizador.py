"""
Router del Cotizador — Editor de cotizaciones con precios calculados.
Endpoints:
  GET  /cotizador/config                        → configuración global
  PUT  /cotizador/config                        → actualizar configuración
  GET  /cotizador/{id}                          → cotización con todos los campos de precio calculados
  PUT  /cotizador/{id}/meta                     → actualizar metadatos (cliente, RUT, etc.)
  PUT  /cotizador/{id}/items/{item_id}          → actualizar ítem editable
  DELETE /cotizador/{id}/items/{item_id}        → eliminar ítem
  POST /cotizador/{id}/formal                   → generar cotización formal (Excel cliente)
  GET  /cotizador/{id}/formal/download          → descargar Excel formal
"""
import os
import uuid
import asyncio
from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import Optional, List
from pydantic import BaseModel
from database import get_db
from models.models import (
    Cotizacion, ItemCotizacion, ConfiguracionCotizador,
    EstadoCotizacion, User, PartsCache,
)
from auth import get_current_user
from services.pricing_service import calcular_cotizacion
from services.scraper import scrape_parts_batch

router = APIRouter(prefix="/cotizador", tags=["cotizador"])

RESULTS_DIR = "results"
os.makedirs(RESULTS_DIR, exist_ok=True)


# ─── helpers ────────────────────────────────────────────────────────────────

def _get_or_create_config(db: Session) -> ConfiguracionCotizador:
    config = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
    if not config:
        config = ConfiguracionCotizador(id=1)
        db.add(config)
        db.commit()
        db.refresh(config)
    return config


def _config_to_dict(config: ConfiguracionCotizador) -> dict:
    return {
        "tipo_cambio_usd": config.tipo_cambio_usd,
        "costo_shipping_usd_kg": config.costo_shipping_usd_kg,
        "adicionales_shipping_usd": config.adicionales_shipping_usd,
        "costo_agencia_pct": config.costo_agencia_pct,
        "costo_agencia_minimo_clp": config.costo_agencia_minimo_clp,
        "desconsolidado_clp": config.desconsolidado_clp,
        "bodegaje_clp": config.bodegaje_clp,
        "margen_venta_pct": config.margen_venta_pct,
        "plazo_min_default": config.plazo_min_default or 30,
        "plazo_max_default": config.plazo_max_default or 45,
    }


def _item_to_dict(item: ItemCotizacion) -> dict:
    return {
        "id": item.id,
        "item_num": item.item_num,
        "numero_parte": item.numero_parte,
        "descripcion": item.descripcion,
        "marca": item.marca,
        "cantidad": item.cantidad,
        "precio_unit_cotizacion": item.precio_unit_cotizacion,
        "peso_unit_lbs": item.peso_unit_lbs,
        "margen_pct": item.margen_pct,
        "plazo": item.plazo,
        "plazo_entrega_min": item.plazo_entrega_min,
        "plazo_entrega_max": item.plazo_entrega_max,
        "precio_finning": item.precio_finning,
        # Datos scraping CAT
        "nombre_cat": item.nombre_cat,
        "precio_cat": item.precio_cat,
        "url_cat": item.url_cat,
        "encontrado": item.encontrado,
    }


def _check_ownership(cotizacion_id: int, user: User, db: Session) -> Cotizacion:
    c = db.query(Cotizacion).filter(
        Cotizacion.id == cotizacion_id,
        Cotizacion.user_id == user.id
    ).first()
    if not c:
        raise HTTPException(status_code=404, detail="Cotización no encontrada")
    return c


# ─── Schemas ────────────────────────────────────────────────────────────────

class ConfigUpdate(BaseModel):
    tipo_cambio_usd: Optional[float] = None
    costo_shipping_usd_kg: Optional[float] = None
    adicionales_shipping_usd: Optional[float] = None
    costo_agencia_pct: Optional[float] = None
    costo_agencia_minimo_clp: Optional[float] = None
    desconsolidado_clp: Optional[float] = None
    bodegaje_clp: Optional[float] = None
    margen_venta_pct: Optional[float] = None
    plazo_min_default: Optional[int] = None
    plazo_max_default: Optional[int] = None


class MetaUpdate(BaseModel):
    cliente: Optional[str] = None
    rut_cliente: Optional[str] = None
    contacto_cliente: Optional[str] = None
    email_cliente: Optional[str] = None
    telefono_cliente: Optional[str] = None
    direccion_cliente: Optional[str] = None
    referencia: Optional[str] = None


class ItemUpdate(BaseModel):
    cantidad: Optional[float] = None
    precio_unit_cotizacion: Optional[float] = None
    peso_unit_lbs: Optional[float] = None
    margen_pct: Optional[float] = None
    plazo: Optional[str] = None
    precio_finning: Optional[float] = None
    descripcion: Optional[str] = None
    marca: Optional[str] = None


# ─── Endpoints ──────────────────────────────────────────────────────────────

@router.get("/config")
def get_config(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    config = _get_or_create_config(db)
    return _config_to_dict(config)


@router.put("/config")
def update_config(
    body: ConfigUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    config = _get_or_create_config(db)
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(config, field, val)
    db.commit()
    db.refresh(config)
    return _config_to_dict(config)


@router.get("/{cotizacion_id}")
def get_editor(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    c = _check_ownership(cotizacion_id, current_user, db)
    config = _get_or_create_config(db)
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
        .order_by(ItemCotizacion.item_num)
        .all()
    )

    items_raw = [_item_to_dict(i) for i in items]
    resultado = calcular_cotizacion(items_raw, _config_to_dict(config))

    return {
        "id": c.id,
        "numero": c.numero,
        "cliente": c.cliente,
        "rut_cliente": c.rut_cliente,
        "contacto_cliente": c.contacto_cliente,
        "email_cliente": c.email_cliente,
        "telefono_cliente": c.telefono_cliente,
        "direccion_cliente": c.direccion_cliente,
        "referencia": c.referencia,
        "estado": c.estado,
        "es_formal": c.es_formal,
        "version_count": db.query(CotizacionVersion).filter(CotizacionVersion.cotizacion_id == cotizacion_id).count(),
        "created_at": c.created_at,
        "terminos_condiciones": c.terminos_condiciones,
        "fase_comercial": c.fase_comercial,
        "config": _config_to_dict(config),
        "items": resultado["items"],
        "totales": resultado["totales"],
    }


@router.put("/{cotizacion_id}/meta")
def update_meta(
    cotizacion_id: int,
    body: MetaUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    c = _check_ownership(cotizacion_id, current_user, db)
    for field, val in body.model_dump(exclude_none=True).items():
        setattr(c, field, val)
    db.commit()
    return {"ok": True}


@router.put("/{cotizacion_id}/items/{item_id}")
def update_item(
    cotizacion_id: int,
    item_id: int,
    body: ItemUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _check_ownership(cotizacion_id, current_user, db)
    item = db.query(ItemCotizacion).filter(
        ItemCotizacion.id == item_id,
        ItemCotizacion.cotizacion_id == cotizacion_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")

    updates = body.model_dump(exclude_none=True)
    for field, val in updates.items():
        setattr(item, field, val)
    db.commit()
    return {"ok": True}


@router.delete("/{cotizacion_id}/items/{item_id}")
def delete_item(
    cotizacion_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _check_ownership(cotizacion_id, current_user, db)
    item = db.query(ItemCotizacion).filter(
        ItemCotizacion.id == item_id,
        ItemCotizacion.cotizacion_id == cotizacion_id,
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Ítem no encontrado")
    db.delete(item)

    # Actualizar total_items en cotización
    c = db.query(Cotizacion).filter(Cotizacion.id == cotizacion_id).first()
    if c:
        c.total_items = db.query(ItemCotizacion).filter(
            ItemCotizacion.cotizacion_id == cotizacion_id
        ).count() - 1  # -1 porque aún no se flush
    db.commit()
    return {"ok": True}


@router.post("/{cotizacion_id}/formal")
def generar_formal(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Genera el Excel de Cotización Cliente (vista formal) y marca como formal."""
    c = _check_ownership(cotizacion_id, current_user, db)
    config = _get_or_create_config(db)
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
        .order_by(ItemCotizacion.item_num)
        .all()
    )

    items_raw = [_item_to_dict(i) for i in items]
    resultado = calcular_cotizacion(items_raw, _config_to_dict(config))

    # Generar Excel formal
    filename = f"formal_{cotizacion_id}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(RESULTS_DIR, filename)
    _generar_excel_formal(c, resultado, filepath, _config_to_dict(config))

    c.archivo_formal = filename
    c.es_formal = 1
    db.commit()

    return {"ok": True, "filename": filename}


@router.get("/{cotizacion_id}/formal/download")
def download_formal(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Regenera siempre el Excel formal con el formato mas reciente."""
    c = _check_ownership(cotizacion_id, current_user, db)
    config = _get_or_create_config(db)
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
        .order_by(ItemCotizacion.item_num)
        .all()
    )
    items_raw = [_item_to_dict(i) for i in items]
    resultado = calcular_cotizacion(items_raw, _config_to_dict(config))
    filename = f"formal_{cotizacion_id}_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(RESULTS_DIR, filename)
    _generar_excel_formal(c, resultado, filepath, _config_to_dict(config))
    c.archivo_formal = filename
    c.es_formal = 1
    db.commit()
    return FileResponse(
        filepath,
        filename=f"CotizacionFormal-{c.numero}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── Generador Excel Formal ─────────────────────────────────────────────────

def _generar_excel_formal(cotizacion: Cotizacion, resultado: dict, output_path: str, config: dict):
    """
    Genera Excel de Cotización Formal — formato exacto CotizacionFormal de referencia.
    Hoja: "Cotizacion Maquinaria"
    6 columnas: Repuesto | Marca | Número de parte | Cantidad | Precio Unit. | TOTAL
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import date, timedelta

    # ── Colores ──
    ORNG   = "E87120"   # naranja labels / header items
    WHITE  = "FFFFFF"
    LGRAY  = "F2F2F2"   # fila par items
    # sin fondo = None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotizacion Maquinaria"
    ws.sheet_view.showGridLines = False

    # ── Anchos de columna (A=37 B=22 C=23 D=18 E=13 F=16) ──
    for col, w in zip("ABCDEFG", [37, 22, 23, 18, 13, 16, 18]):
        ws.column_dimensions[col].width = w

    def st(cell, bold=False, color="000000", bg=None, align="left", size=9, wrap=False):
        cell.font = Font(bold=bold, color=color, size=size, name="Calibri")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

    def label(cell, text, size=9):
        """Label naranja bold"""
        cell.value = text
        st(cell, bold=True, color=ORNG, size=size)

    def value(cell, text):
        """Valor normal"""
        cell.value = text
        st(cell, size=9)

    def spacer(row_num, h=6):
        ws.row_dimensions[row_num].height = h

    today = date.today()
    valid_until = today + timedelta(days=15)

    fases = {
        "ingresada": "Propuesta", "validada": "Validada",
        "semi_validada": "Semi-Validada", "pre_embarcado": "Pre-Embarcado",
        "embarcado": "Embarcado", "pagado": "Pagado",
    }
    estado = fases.get(cotizacion.fase_comercial or "ingresada", "Propuesta")

    items_calc = resultado.get("items", [])
    marcas = [i.get("marca", "") for i in items_calc if i.get("marca")]
    marca_display = marcas[0] if marcas else ""

    fecha_str = today.strftime("%d/%m/%Y")

    # ── Filas 1-4: Header empresa ──
    for r in [1, 2, 3, 4]:
        ws.row_dimensions[r].height = 14

    # A1:B4 — área logo (fondo blanco, vacío)
    ws.merge_cells("A1:B4")
    ws["A1"].fill = PatternFill("solid", fgColor=WHITE)

    # D1:F1 — Grupo AM SpA
    ws.merge_cells("D1:G1")
    c1 = ws["D1"]
    c1.value = "Grupo AM SpA"
    st(c1, bold=True, color="000000", align="right", size=11)

    ws.merge_cells("D2:G2")
    c2 = ws["D2"]
    c2.value = "77.977.813-4"
    st(c2, color="000000", align="right", size=10)

    ws.merge_cells("D3:G3")
    c3 = ws["D3"]
    c3.value = "Alonso de Cordova 5870 Of 413, Santiago"
    st(c3, color="000000", align="right", size=10)

    # ── Fila 5: COTIZACIÓN Nº ──
    ws.row_dimensions[5].height = 22
    ws.merge_cells("A5:G5")
    c5 = ws["A5"]
    c5.value = f"COTIZACIÓN Nº : {cotizacion.numero}"
    st(c5, bold=True, color=ORNG, align="center", size=14)

    # ── Fila 6: espaciador ──
    spacer(6)

    # ── Filas 7-11: Info sección ──
    for r in range(7, 12):
        ws.row_dimensions[r].height = 14

    # Fila 7
    label(ws["A7"], "Tipo Cotización:")
    value(ws["B7"], "Cotización Importación de Repuestos")
    label(ws["D7"], "Estado:")
    value(ws["E7"], estado)

    # Fila 8
    label(ws["A8"], "Fecha Solicitud:")
    value(ws["B8"], fecha_str)
    label(ws["D8"], "Forma Pago:")
    value(ws["E8"], "30 días contra factura")

    # Fila 9
    label(ws["A9"], "Marca:")
    value(ws["B9"], marca_display)
    label(ws["D9"], "Modelo:")
    # E9 vacío

    # Fila 10
    label(ws["A10"], "VIN:")
    label(ws["D10"], "Año:")

    # Fila 11 — Descripción (valor merged B11:F11)
    label(ws["A11"], "Descripción:")
    ws.merge_cells("B11:G11")
    desc_cell = ws["B11"]
    desc_cell.value = cotizacion.referencia or "Por solicitud del cliente se presenta cotización por importación de repuestos"
    st(desc_cell, size=9, wrap=True)

    # ── Fila 12: espaciador ──
    spacer(12)

    # ── Fila 13: Headers cliente ──
    ws.row_dimensions[13].height = 14
    label(ws["A13"], "Datos Cliente", size=10)
    ws.merge_cells("D13:G13")
    label(ws["D13"], "Datos para Orden de Compra", size=10)

    # ── Filas 14-18: Datos cliente ──
    for r in range(14, 19):
        ws.row_dimensions[r].height = 14

    # Fila 14
    label(ws["A14"], "Cliente:")
    value(ws["B14"], cotizacion.cliente or "")
    label(ws["D14"], "Razón social:")
    ws.merge_cells("E14:G14")
    value(ws["E14"], "GRUPO AM SPA")

    # Fila 15
    label(ws["A15"], "Rut:")
    value(ws["B15"], cotizacion.rut_cliente or "")
    label(ws["D15"], "Rut:")
    ws.merge_cells("E15:G15")
    value(ws["E15"], "77.977.813-4")

    # Fila 16
    label(ws["A16"], "Atención:")
    value(ws["B16"], cotizacion.contacto_cliente or "")
    label(ws["D16"], "Dirección:")
    ws.merge_cells("E16:G16")
    value(ws["E16"], "Alonso de Cordova 5870 of 413, Santiago")

    # Fila 17
    label(ws["A17"], "Servicio:")
    value(ws["B17"], "Importación de Repuestos")
    label(ws["D17"], "Giro:")
    ws.merge_cells("E17:G17")
    value(ws["E17"], "Venta de repuestos")

    # Fila 18
    label(ws["A18"], "Comentario:")
    # B18 vacío
    label(ws["D18"], "Correo:")
    ws.merge_cells("E18:G18")
    value(ws["E18"], "aibaceta@gruporamaq.cl")

    # ── Fila 19: espaciador ──
    spacer(19)

    # ── Fila 20: Header tabla items ──
    ws.row_dimensions[20].height = 18
    headers = [("A", "Repuesto", "left"), ("B", "Marca", "left"),
               ("C", "Número de parte", "left"), ("D", "Cantidad", "center"),
               ("E", "Precio Unit.", "center"), ("F", "TOTAL", "center"),
               ("G", "Plazo de Entrega", "center")]
    for col, txt, aln in headers:
        cell = ws[f"{col}20"]
        cell.value = txt
        st(cell, bold=True, color=WHITE, bg=ORNG, align=aln, size=9)

    # ── Filas de items ──
    data_start = 21
    for i, item in enumerate(items_calc):
        row = data_start + i
        ws.row_dimensions[row].height = 14
        bg = LGRAY if i % 2 == 0 else WHITE

        # A: descripción
        ca = ws[f"A{row}"]
        ca.value = item.get("descripcion") or ""
        st(ca, bg=bg, size=9)

        # B: marca
        cb = ws[f"B{row}"]
        cb.value = item.get("marca") or ""
        st(cb, bg=bg, size=9)

        # C: número de parte
        cc = ws[f"C{row}"]
        cc.value = item.get("numero_parte") or ""
        st(cc, bg=bg, size=9)

        # D: cantidad (center)
        cd = ws[f"D{row}"]
        cd.value = item.get("cantidad") or 0
        st(cd, bg=bg, align="center", size=9)

        # E: precio unit (right, number format)
        ce = ws[f"E{row}"]
        precio = item.get("precio_venta_clp") or 0
        ce.value = int(round(precio)) if precio else 0
        st(ce, bg=bg, align="right", size=9)
        ce.number_format = '#,##0'

        # F: total (right, number format)
        cf = ws[f"F{row}"]
        total = item.get("total_venta_clp") or 0
        cf.value = int(round(total)) if total else 0
        st(cf, bg=bg, align="right", size=9)
        cf.number_format = '#,##0'

        # G: plazo de entrega — fallback a defaults de config
        cg = ws[f"G{row}"]
        _mn = item.get("plazo_entrega_min") or config.get("plazo_min_default")
        _mx = item.get("plazo_entrega_max") or config.get("plazo_max_default")
        if _mn and _mx:
            _plazo_txt = f"de {_mn} a {_mx} días hábiles"
        elif _mn:
            _plazo_txt = f"{_mn} días hábiles"
        elif _mx:
            _plazo_txt = f"{_mx} días hábiles"
        else:
            _plazo_txt = item.get("plazo") or "A confirmar"
        cg.value = _plazo_txt
        st(cg, bg=bg, align="center", size=9)

    # ── Totales (debajo de items) ──
    totales = resultado.get("totales", {})
    tr = data_start + len(items_calc)  # fila vacía entre items y totales

    # Fila vacía
    ws.row_dimensions[tr].height = 6

    # TOTAL NETO
    r1 = tr + 1
    ws.row_dimensions[r1].height = 14
    e1 = ws[f"E{r1}"]
    e1.value = "TOTAL NETO."
    st(e1, bold=True, align="right", size=9)
    f1 = ws[f"F{r1}"]
    f1.value = int(round(totales.get("subtotal_neto_clp", 0)))
    st(f1, align="right", size=9)
    f1.number_format = '#,##0'

    # IVA
    r2 = tr + 2
    ws.row_dimensions[r2].height = 14
    e2 = ws[f"E{r2}"]
    e2.value = "IVA."
    st(e2, bold=True, align="right", size=9)
    f2 = ws[f"F{r2}"]
    f2.value = int(round(totales.get("iva_clp", 0)))
    st(f2, align="right", size=9)
    f2.number_format = '#,##0'

    # TOTAL (naranja)
    r3 = tr + 3
    ws.row_dimensions[r3].height = 14
    e3 = ws[f"E{r3}"]
    e3.value = "TOTAL."
    st(e3, bold=True, color=WHITE, bg=ORNG, align="right", size=9)
    f3 = ws[f"F{r3}"]
    f3.value = int(round(totales.get("total_con_iva_clp", 0)))
    st(f3, bold=True, color=WHITE, bg=ORNG, align="right", size=9)
    f3.number_format = '#,##0'

    # ── Condiciones de Servicio ──
    cond_row = tr + 5
    ws.row_dimensions[cond_row].height = 14
    ws.merge_cells(f"A{cond_row}:G{cond_row}")
    label(ws[f"A{cond_row}"], "Condiciones de Servicio", size=10)

    if cotizacion.terminos_condiciones:
        conds = [l.strip() for l in cotizacion.terminos_condiciones.splitlines() if l.strip()]
    else:
        tc = config.get("tipo_cambio_usd", 940)
        plazo_txt = _compute_plazo_text(items_calc, config)
        conds = [
            "1.- Validez de la cotización 30 días corridos desde la fecha de emisión.",
            f"2.- {plazo_txt}",
            "3.- Precios expresados en CLP. IVA no incluido.",
            "4.- Repuestos genuinos, en su packing original.",
            f"5.- Tipo de cambio referencial: ${tc:,.0f} CLP/USD.",
        ]
    for j, cond in enumerate(conds):
        r = cond_row + 1 + j
        ws.row_dimensions[r].height = 14
        ws.merge_cells(f"A{r}:G{r}")
        value(ws[f"A{r}"], cond)
        ws[f"A{r}"].font = ws[f"A{r}"].font.copy(size=8)

    # ── Datos Bancarios ──
    bank_start = cond_row + len(conds) + 2
    ws.row_dimensions[bank_start].height = 14
    ws.merge_cells(f"A{bank_start}:G{bank_start}")
    label(ws[f"A{bank_start}"], "Datos Bancarios", size=10)

    bank_rows = [
        ("Razón Social:", "Grupo AM SpA", "Banco:", "Banco Santander"),
        ("Rut:", "77.977.813-4", "Tipo Cuenta:", "Cuenta Corriente"),
        ("N° Cuenta:", "96353421", "Mail:", "aibaceta@gruporamaq.cl"),
    ]
    for j, (l1, v1, l2, v2) in enumerate(bank_rows):
        r = bank_start + 1 + j
        ws.row_dimensions[r].height = 14
        label(ws[f"A{r}"], l1)
        value(ws[f"B{r}"], v1)
        label(ws[f"D{r}"], l2)
        value(ws[f"E{r}"], v2)

    wb.save(output_path)


def _compute_plazo_text(items_calc: list, config: dict) -> str:
    """Devuelve la frase de plazo de entrega usando los plazos minimo/maximo
    de los items (con fallback a config.plazo_min_default / plazo_max_default).
    Ejemplo: 'Plazo de entrega de 7 a 10 días.'
    """
    mins = [it.get("plazo_entrega_min") for it in items_calc if it.get("plazo_entrega_min")]
    maxs = [it.get("plazo_entrega_max") for it in items_calc if it.get("plazo_entrega_max")]
    pmin = min(mins) if mins else (config.get("plazo_min_default") or 30)
    pmax = max(maxs) if maxs else (config.get("plazo_max_default") or 45)
    return f"Plazo de entrega de {pmin} a {pmax} días."


DEFAULT_TERMINOS = """• Precios en CLP no incluyen IVA.
• Forma de pago: 50% anticipo, 50% contra entrega.
• Plazo de entrega sujeto a confirmación de stock.
• Garantía: 6 meses por defectos de fabricación."""


class TerminosUpdate(BaseModel):
    terminos: str


@router.put("/{cotizacion_id}/terminos")
def update_terminos(
    cotizacion_id: int,
    body: TerminosUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Guarda los términos y condiciones de la cotización."""
    c = _check_ownership(cotizacion_id, current_user, db)
    c.terminos_condiciones = body.terminos
    db.commit()
    return {"ok": True}


# ── Rutas adicionales (versiones, scraping Finning, PDF, estados) ─────────────

from models.models import CotizacionVersion
import json as _json


@router.get("/{cotizacion_id}/versiones")
def listar_versiones(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lista las versiones guardadas de una cotización."""
    _check_ownership(cotizacion_id, current_user, db)
    vers = db.query(CotizacionVersion).filter(
        CotizacionVersion.cotizacion_id == cotizacion_id
    ).order_by(CotizacionVersion.version_num.desc()).all()
    return [
        {
            "id": v.id,
            "version_num": v.version_num,
            "descripcion": v.descripcion or "",
            "created_at": v.created_at.isoformat() if v.created_at else None,
        }
        for v in vers
    ]


@router.get("/{cotizacion_id}/versiones/{version_num}")
def get_version(
    cotizacion_id: int,
    version_num: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Obtiene el snapshot de una versión específica."""
    _check_ownership(cotizacion_id, current_user, db)
    ver = db.query(CotizacionVersion).filter(
        CotizacionVersion.cotizacion_id == cotizacion_id,
        CotizacionVersion.version_num == version_num,
    ).first()
    if not ver:
        raise HTTPException(404, "Versión no encontrada")
    try:
        return _json.loads(ver.snapshot_json)
    except Exception:
        return {"raw": ver.snapshot_json}


@router.post("/{cotizacion_id}/versiones/{version_num}/restaurar")
def restaurar_version(
    cotizacion_id: int,
    version_num: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Restaura el estado de una cotización a una versión anterior."""
    cotizacion = _check_ownership(cotizacion_id, current_user, db)
    ver = db.query(CotizacionVersion).filter(
        CotizacionVersion.cotizacion_id == cotizacion_id,
        CotizacionVersion.version_num == version_num,
    ).first()
    if not ver:
        raise HTTPException(404, "Versión no encontrada")
    try:
        snap = _json.loads(ver.snapshot_json)
        # Restore items from snapshot
        items_snap = snap.get("items", [])
        existing = {i.id: i for i in cotizacion.items}
        for idata in items_snap:
            item = existing.get(idata.get("id"))
            if item:
                for k in ["cantidad", "precio_unit_cotizacion", "peso_unit_lbs",
                          "margen_pct", "plazo_entrega_min", "plazo_entrega_max",
                          "descripcion", "marca", "precio_finning"]:
                    if k in idata:
                        setattr(item, k, idata[k])
        db.commit()
        return {"ok": True, "version": version_num}
    except Exception as e:
        raise HTTPException(500, f"Error al restaurar: {str(e)}")


class ItemEstadoUpdate(BaseModel):
    item_id: int
    estado_item: str


@router.put("/{cotizacion_id}/items/estados")
def actualizar_estados_items(
    cotizacion_id: int,
    body: List[ItemEstadoUpdate],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Actualiza el estado de múltiples items a la vez."""
    _check_ownership(cotizacion_id, current_user, db)
    updated = 0
    for upd in body:
        item = db.query(ItemCotizacion).filter(
            ItemCotizacion.id == upd.item_id,
            ItemCotizacion.cotizacion_id == cotizacion_id,
        ).first()
        if item:
            item.estado_item = upd.estado_item
            updated += 1
    db.commit()
    return {"ok": True, "updated": updated}


async def _scrape_finning_for_cotizacion(cotizacion_id: int, db_url: str):
    """Background task: scrapea precios Finning (parts.cat.com/es/finningchile)
    para los items de una cotización ya existente en BD.

    Usa el mismo scraper que el upload Excel (services.scraper.scrape_parts_batch),
    con cache (PartsCache) para evitar duplicados.
    """
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker

    engine = create_engine(db_url, pool_pre_ping=True)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        cot = db.query(Cotizacion).filter(Cotizacion.id == cotizacion_id).first()
        if not cot:
            return

        cot.estado = EstadoCotizacion.PROCESANDO
        cot.items_procesados = 0
        cot.items_encontrados = 0
        db.commit()

        items = (
            db.query(ItemCotizacion)
            .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
            .all()
        )
        part_numbers = [it.numero_parte for it in items if it.numero_parte]

        # Cache check
        cached = {}
        for np in part_numbers:
            entry = db.query(PartsCache).filter(PartsCache.numero_parte == np).first()
            if entry:
                cached[np] = entry

        to_scrape = [np for np in part_numbers if np not in cached]

        scraped = {}
        if to_scrape:
            results = await scrape_parts_batch(to_scrape)
            for r in results:
                np = r["numero_parte"]
                scraped[np] = r
                entry = db.query(PartsCache).filter(PartsCache.numero_parte == np).first()
                if not entry:
                    entry = PartsCache(numero_parte=np)
                    db.add(entry)
                entry.nombre_cat = r.get("nombre_cat")
                entry.precio_cat = r.get("precio_cat")
                entry.moneda_cat = r.get("moneda_cat", "CLP")
                entry.retiro_estimado = r.get("retiro_estimado")
                entry.url_cat = r.get("url_cat")
                entry.imagen_url = r.get("imagen_url")
                entry.encontrado = r.get("encontrado", 0)
                db.commit()

        # Update items with results
        encontrados = 0
        for item in items:
            np = item.numero_parte
            data = None
            if np in cached:
                c = cached[np]
                data = {
                    "nombre_cat": c.nombre_cat,
                    "precio_cat": c.precio_cat,
                    "moneda_cat": c.moneda_cat,
                    "retiro_estimado": c.retiro_estimado,
                    "url_cat": c.url_cat,
                    "imagen_url": c.imagen_url,
                    "encontrado": c.encontrado,
                }
            elif np in scraped:
                data = scraped[np]
            if data:
                item.nombre_cat = data.get("nombre_cat")
                item.precio_cat = data.get("precio_cat")
                item.moneda_cat = data.get("moneda_cat", "CLP")
                item.retiro_estimado = data.get("retiro_estimado")
                item.url_cat = data.get("url_cat")
                item.imagen_url = data.get("imagen_url")
                item.encontrado = data.get("encontrado", 0)
                if data.get("encontrado"):
                    encontrados += 1
            cot.items_procesados = (cot.items_procesados or 0) + 1
            db.commit()

        cot.items_encontrados = encontrados
        cot.estado = EstadoCotizacion.COMPLETADO
        cot.total_items = len(items)
        # Limpiar error_msg residual de corridas anteriores si esta fue exitosa
        if encontrados > 0:
            cot.error_msg = None
        elif len(items) > 0:
            # 0 encontrados de N items procesados: probablemente Akamai bloqueó
            # (parts.cat.com tiene anti-bot Akamai). Marcamos ERROR con mensaje
            # visible en la UI.
            cot.estado = EstadoCotizacion.ERROR
            cot.error_msg = (
                "El portal Finning bloqueó la búsqueda automática (anti-bot Akamai). "
                "Reintentá más tarde o ingresá los precios manualmente."
            )
        db.commit()
    except Exception as e:
        try:
            cot = db.query(Cotizacion).filter(Cotizacion.id == cotizacion_id).first()
            if cot:
                cot.estado = EstadoCotizacion.ERROR
                cot.error_msg = str(e)[:500]
                db.commit()
        except Exception:
            pass
    finally:
        db.close()


@router.post("/{cotizacion_id}/buscar-finning")
def buscar_finning(
    cotizacion_id: int,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Dispara búsqueda real de precios Finning para los items de una cotización.

    Usa el mismo scraper que el upload Excel inicial (Playwright + parts.cat.com/es/finningchile).
    El scraping corre en background; el frontend hace polling vía /finning-progress.
    """
    cotizacion = _check_ownership(cotizacion_id, current_user, db)
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
        .all()
    )
    if not items:
        raise HTTPException(400, "La cotización no tiene items")

    cotizacion.estado = EstadoCotizacion.PROCESANDO
    cotizacion.items_procesados = 0
    cotizacion.items_encontrados = 0
    cotizacion.total_items = len(items)
    db.commit()

    from config import settings
    background_tasks.add_task(
        asyncio.run,
        _scrape_finning_for_cotizacion(cotizacion_id, settings.DATABASE_URL),
    )
    return {
        "ok": True,
        "message": f"Búsqueda iniciada para {len(items)} items.",
        "total_items": len(items),
    }


@router.get("/{cotizacion_id}/finning-progress")
def finning_progress(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Retorna el progreso del scraping Finning.

    Schema esperado por el frontend (CotizadorEditor.tsx):
      - completado: bool  (true cuando estado != PROCESANDO/ESPERANDO_AGENTE)
      - total:      int
      - con_precio: int
      - sin_precio: int
      - pct:        int (0-100)
      - estado:     enum string
      - error_msg:  str | null
    """
    cotizacion = _check_ownership(cotizacion_id, current_user, db)

    # Contar items con precio real (precio_cat NOT NULL/0)
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
        .all()
    )
    total = len(items)
    con_precio = sum(1 for i in items if i.precio_cat)
    sin_precio = total - con_precio
    pct = int((con_precio / total) * 100) if total > 0 else 0

    estado_val = cotizacion.estado
    estado_str = estado_val.value if hasattr(estado_val, "value") else str(estado_val)
    en_proceso = estado_str in ("procesando", "esperando_agente")
    completado = not en_proceso

    return {
        "completado": completado,
        "total": total,
        "con_precio": con_precio,
        "sin_precio": sin_precio,
        "pct": pct,
        "estado": estado_str,
        "error_msg": cotizacion.error_msg,
        # backward compat
        "total_items": cotizacion.total_items or 0,
        "items_procesados": cotizacion.items_procesados or 0,
        "items_encontrados": cotizacion.items_encontrados or 0,
    }


def _generar_pdf_formal(cotizacion, resultado: dict, output_path: str, config: dict):
    """Genera PDF de cotización — formato Grupo AM SpA (Cotizacion-4.pdf)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.colors import HexColor, white, black
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.pdfgen import canvas as pdfcanvas
    from datetime import datetime

    AZUL  = HexColor("#1A3A5C")
    AZUL_M= HexColor("#2E5984")
    NARJ  = HexColor("#EA6C0D")
    GRIS  = HexColor("#F5F5F5")
    GRIS2 = HexColor("#E8E8E8")
    BORDE = HexColor("#CCCCCC")

    W_PAGE, H_PAGE = A4
    MARGIN = 1.5 * cm
    W = W_PAGE - 2 * MARGIN

    def ps(name, size=9, bold=False, color=black, align=TA_LEFT, leading=None):
        return ParagraphStyle(
            name, fontSize=size,
            fontName="Helvetica-Bold" if bold else "Helvetica",
            textColor=color, alignment=align,
            leading=leading or size * 1.4,
        )

    def clp(v):
        try:
            return f"${float(v):,.0f}".replace(",", ".")
        except:
            return "—"

    class FooterCanvas(pdfcanvas.Canvas):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)
            self._saved = []
        def showPage(self):
            self._saved.append(dict(self.__dict__))
            self._startPage()
        def save(self):
            total = len(self._saved)
            for state in self._saved:
                self.__dict__.update(state)
                self._footer(total)
                super().showPage()
            super().save()
        def _footer(self, total):
            self.saveState()
            self.setStrokeColor(BORDE)
            self.setLineWidth(0.5)
            self.line(MARGIN, 1.4 * cm, W_PAGE - MARGIN, 1.4 * cm)
            self.setFont("Helvetica", 7.5)
            self.setFillColor(HexColor("#888888"))
            self.drawString(MARGIN, 1.0 * cm,
                "Grupo AM SpA  |  RUT 77.977.813-4  |  aibaceta@gruporamaq.cl")
            self.drawRightString(W_PAGE - MARGIN, 1.0 * cm,
                f"Página {self._pageNumber} de {total}")
            self.restoreState()

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        leftMargin=MARGIN, rightMargin=MARGIN,
        topMargin=MARGIN, bottomMargin=2.0 * cm,
    )
    story = []

    fecha_str = cotizacion.created_at.strftime("%d/%m/%Y") if cotizacion.created_at else datetime.now().strftime("%d/%m/%Y")
    fases = {
        "ingresada": "Propuesta", "validada": "Validada",
        "semi_validada": "Semi-Validada", "pre_embarcado": "Pre-Embarcado",
        "embarcado": "Embarcado", "pagado": "Pagado",
    }
    estado = fases.get(cotizacion.fase_comercial or "ingresada", "Propuesta")
    items_calc = resultado.get("items", [])
    marcas = [i.get("marca", "") for i in items_calc if i.get("marca")]
    marca_display = marcas[0] if marcas else "—"
    desc_display = cotizacion.referencia or "Por solicitud del cliente se presenta cotización por importación de repuestos"

    # ── 1. Header ──────────────────────────────────────────────────────────
    left_txt = (
        f'<font size="18"><b><font color="#1A3A5C">Grupo AM SpA</font></b></font><br/>'
        f'<font size="8" color="#555555">RUT: 77.977.813-4</font><br/>'
        f'<font size="8" color="#555555">Alonso de Córdoba 5870 Of 413, Santiago</font>'
    )
    right_txt = (
        f'<b><font size="13" color="#1A3A5C">COTIZACIÓN Nº {cotizacion.numero}</font></b><br/>'
        f'<font size="9" color="#333333">Fecha: {fecha_str}</font><br/>'
        f'<font size="9" color="#333333">Estado: {estado}</font>'
    )
    t_hdr = Table(
        [[Paragraph(left_txt, ps("hl")), Paragraph(right_txt, ps("hr", align=TA_RIGHT))]],
        colWidths=[W * 0.55, W * 0.45],
    )
    t_hdr.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story.append(t_hdr)
    story.append(Spacer(1, 0.2 * cm))
    story.append(HRFlowable(width="100%", thickness=2.5, color=NARJ, spaceAfter=2))
    story.append(Spacer(1, 0.25 * cm))

    # ── 2. Info grid (3 filas × 4 cols) ───────────────────────────────────
    lw = 2.8 * cm   # label width per side
    vw = 6.2 * cm   # value width per side  (lw+vw = 9.0cm = W/2)

    def il(t): return Paragraph(f"<b>{t}</b>", ps(f"il{t}", size=8, bold=True, color=white))
    def iv(t): return Paragraph(str(t or "—"), ps(f"iv{t}", size=8))

    info_data = [
        [il("Tipo Cotización:"),  iv("Cotización Importación de Repuestos"), il("Forma de Pago:"), iv("30 días contra factura")],
        [il("Fecha Solicitud:"),  iv(fecha_str),                             il("Marca:"),         iv(marca_display)],
        [il("Descripción:"),      iv(desc_display),                          il("Referencia:"),    iv(cotizacion.referencia or "—")],
    ]
    t_info = Table(info_data, colWidths=[lw, vw, lw, vw])
    t_info.setStyle(TableStyle([
        ("BACKGROUND",  (0, 0), (0, -1), AZUL_M),
        ("BACKGROUND",  (2, 0), (2, -1), AZUL_M),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",(0, 0), (-1, -1), 6),
        ("BOX",         (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID",   (0, 0), (-1, -1), 0.3, BORDE),
    ]))
    story.append(t_info)
    story.append(Spacer(1, 0.35 * cm))

    # ── 3. Datos Cliente + Datos para OC ──────────────────────────────────
    cw2 = W / 2   # exactly half of W so both columns align with other tables

    def clbl(t): return Paragraph(f"<b>{t}</b>", ps(f"clbl{t}", size=8, bold=True, color=white))
    def cval(t): return Paragraph(str(t or "—"), ps(f"cval{t}", size=8))

    lbl_w = 2.4 * cm
    val_w = cw2 - lbl_w   # 9.0 - 2.4 = 6.6 cm

    # Left (Datos Cliente)
    left_data = [
        [Paragraph("<b>Datos Cliente</b>", ps("dch", size=9, bold=True, color=white)), ""],
        [clbl("Cliente:"),   cval(cotizacion.cliente)],
        [clbl("RUT:"),       cval(cotizacion.rut_cliente)],
        [clbl("Atención:"),  cval(cotizacion.contacto_cliente)],
        [clbl("Servicio:"),  cval("Importación de Repuestos")],
        [clbl("Correo:"),    cval(cotizacion.email_cliente)],
    ]
    t_left = Table(left_data, colWidths=[lbl_w, val_w])
    left_style = [
        ("SPAN",        (0, 0), (1, 0)),
        ("BACKGROUND",  (0, 0), (1, 0), AZUL),
        ("BACKGROUND",  (0, 1), (0, -1), AZUL_M),
        ("BACKGROUND",  (1, 1), (1, -1), white),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",(0, 0), (-1, -1), 6),
        ("BOX",         (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID",   (0, 0), (-1, -1), 0.3, BORDE),
    ]
    for i in range(1, len(left_data)):
        if i % 2 == 0:
            left_style.append(("BACKGROUND", (1, i), (1, i), GRIS))
    t_left.setStyle(TableStyle(left_style))

    # Right (Datos para OC)
    right_data = [
        [Paragraph("<b>Datos para Orden de Compra</b>", ps("och", size=9, bold=True, color=white)), ""],
        [clbl("Razón Social:"), cval("GRUPO AM SPA")],
        [clbl("RUT:"),          cval("77.977.813-4")],
        [clbl("Dirección:"),    cval("Alonso de Córdoba 5870 Of 413, Santiago")],
        [clbl("Giro:"),         cval("Venta de repuestos")],
        [clbl("Correo OC:"),    cval("aibaceta@gruporamaq.cl")],
    ]
    t_right = Table(right_data, colWidths=[lbl_w, val_w])
    right_style = [
        ("SPAN",        (0, 0), (1, 0)),
        ("BACKGROUND",  (0, 0), (1, 0), AZUL),
        ("BACKGROUND",  (0, 1), (0, -1), AZUL_M),
        ("BACKGROUND",  (1, 1), (1, -1), white),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING",(0, 0), (-1, -1), 6),
        ("BOX",         (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID",   (0, 0), (-1, -1), 0.3, BORDE),
    ]
    for i in range(1, len(right_data)):
        if i % 2 == 0:
            right_style.append(("BACKGROUND", (1, i), (1, i), GRIS))
    t_right.setStyle(TableStyle(right_style))

    t_clients = Table([[t_left, t_right]], colWidths=[cw2, cw2])
    t_clients.setStyle(TableStyle([
        ("VALIGN",       (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING",  (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING",   (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 0),
    ]))
    story.append(t_clients)
    story.append(Spacer(1, 0.35 * cm))

    # ── 4. Items table ────────────────────────────────────────────────────
    col_ws = [0.7*cm, 5.5*cm, 1.8*cm, 2.0*cm, 0.8*cm, 2.2*cm, 2.0*cm, 3.0*cm]  # sum = 18.0cm = W
    aligns = [TA_CENTER, TA_LEFT, TA_CENTER, TA_CENTER, TA_CENTER, TA_RIGHT, TA_RIGHT, TA_CENTER]
    headers = ["N°", "Repuesto / Descripción", "Marca", "N° Parte", "Qty", "Precio Unit.", "Total", "Plazo Entrega"]

    hdr_row = [
        Paragraph(f"<b>{h}</b>", ps(f"th{i}", size=8, bold=True, color=white, align=aligns[i]))
        for i, h in enumerate(headers)
    ]
    rows = [hdr_row]
    for i, item in enumerate(items_calc):
        bg = GRIS if i % 2 == 0 else white
        n_parte = str(item.get("numero_parte") or "")
        # Plazo de entrega — fallback a defaults de config
        _mn = item.get("plazo_entrega_min") or config.get("plazo_min_default")
        _mx = item.get("plazo_entrega_max") or config.get("plazo_max_default")
        if _mn and _mx:
            _plazo_txt = f"de {_mn} a {_mx} días hábiles"
        elif _mn:
            _plazo_txt = f"{_mn} días hábiles"
        elif _mx:
            _plazo_txt = f"{_mx} días hábiles"
        else:
            _plazo_txt = item.get("plazo") or "A confirmar"
        row_data = [
            Paragraph(str(item.get("item_num") or i+1),  ps(f"r{i}0", size=8, align=TA_CENTER)),
            Paragraph(str(item.get("descripcion") or ""), ps(f"r{i}1", size=8)),
            Paragraph(str(item.get("marca") or ""),       ps(f"r{i}2", size=8, align=TA_CENTER)),
            Paragraph(f"<b>{n_parte}</b>",                ps(f"r{i}3", size=8, bold=True, align=TA_CENTER, color=AZUL)),
            Paragraph(str(item.get("cantidad") or ""),    ps(f"r{i}4", size=8, align=TA_CENTER)),
            Paragraph(clp(item.get("precio_venta_clp")), ps(f"r{i}5", size=8, align=TA_RIGHT)),
            Paragraph(clp(item.get("total_venta_clp")),  ps(f"r{i}6", size=8, align=TA_RIGHT)),
            Paragraph(_plazo_txt,                         ps(f"r{i}7", size=8, align=TA_CENTER)),
        ]
        rows.append(row_data)

    t_items = Table(rows, colWidths=col_ws, repeatRows=1)
    item_style = [
        ("BACKGROUND", (0, 0), (-1, 0), AZUL),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0),(-1, -1), 4),
        ("LEFTPADDING",(0, 0), (-1, -1), 4),
        ("RIGHTPADDING",(0,0), (-1, -1), 4),
        ("BOX",        (0, 0), (-1, -1), 0.5, BORDE),
        ("INNERGRID",  (0, 0), (-1, -1), 0.3, BORDE),
    ]
    for i in range(1, len(rows)):
        bg = GRIS if (i - 1) % 2 == 0 else white
        item_style.append(("BACKGROUND", (0, i), (-1, i), bg))
    t_items.setStyle(TableStyle(item_style))
    story.append(t_items)
    story.append(Spacer(1, 0.2 * cm))

    # ── 5. Totales ────────────────────────────────────────────────────────
    totales = resultado.get("totales", {})
    tot_rows = [
        ("TOTAL NETO", totales.get("subtotal_neto_clp", 0), False),
        ("IVA (19%)",  totales.get("iva_clp", 0),           False),
        ("TOTAL",      totales.get("total_con_iva_clp", 0),  True),
    ]
    for label, val, bold in tot_rows:
        td = Table(
            [[Paragraph(f"<b>{label}</b>",
                        ps(f"tl{label}", size=10 if bold else 9, bold=True,
                           color=AZUL if not bold else white, align=TA_RIGHT)),
              Paragraph(f"<b>{clp(val)}</b>",
                        ps(f"tv{label}", size=10 if bold else 9, bold=True,
                           color=NARJ if not bold else white, align=TA_RIGHT))]],
            colWidths=[W - 4.5*cm, 4.5*cm],
        )
        td.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), AZUL if bold else GRIS2),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5 if bold else 4),
            ("BOTTOMPADDING",(0,0),(-1,-1),  5 if bold else 4),
            ("LEFTPADDING",(0, 0), (-1, -1), 8),
            ("RIGHTPADDING",(0,0), (-1, -1), 8),
            ("BOX",        (0, 0), (-1, -1), 0.5, BORDE),
        ]))
        story.append(td)

    story.append(Spacer(1, 0.4 * cm))

    # ── 6. Condiciones de Servicio ────────────────────────────────────────
    t_ch = Table([[Paragraph("<b>Condiciones de Servicio</b>",
                             ps("csh", size=9, bold=True, color=white))]], colWidths=[W])
    t_ch.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), AZUL),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(t_ch)

    if cotizacion.terminos_condiciones:
        conds = [l.strip() for l in cotizacion.terminos_condiciones.splitlines() if l.strip()]
    else:
        tc = config.get("tipo_cambio_usd", 940)
        plazo_txt = _compute_plazo_text(items_calc, config)
        conds = [
            "1. Validez de la cotización: 30 días corridos desde la fecha de emisión.",
            "2. Precios expresados en CLP. IVA no incluido.",
            "3. Forma de pago: 50% anticipo, saldo contra entrega.",
            f"4. {plazo_txt}",
            "5. Repuestos genuinos en su packing original.",
            f"6. Tipo de cambio referencial: ${tc:,.0f} CLP/USD.",
        ]
    for i, cond in enumerate(conds):
        bg = GRIS if i % 2 == 0 else white
        t = Table([[Paragraph(cond, ps(f"c{i}", size=8))]], colWidths=[W])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), bg),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING",(0,0),(-1,-1), 3),
            ("LEFTPADDING",(0, 0), (-1, -1), 8),
            ("BOX",        (0, 0), (-1, -1), 0.3, BORDE),
        ]))
        story.append(t)

    story.append(Spacer(1, 0.4 * cm))

    # ── 7. Datos Bancarios ────────────────────────────────────────────────
    t_bh = Table([[Paragraph("<b>Datos Bancarios</b>",
                             ps("dbh", size=9, bold=True, color=white))]], colWidths=[W])
    t_bh.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), AZUL),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",(0, 0), (-1, -1), 8),
    ]))
    story.append(t_bh)

    bank_lw = 2.4 * cm
    bank_vw = W / 2 - bank_lw  # 9.0 - 2.4 = 6.6 cm; 2*(2.4+6.6)=18.0cm=W
    bank_rows_data = [
        ("Razón Social:", "Grupo AM SpA",    "Banco:",       "Banco Santander"),
        ("RUT:",          "77.977.813-4",    "Tipo Cuenta:", "Cuenta Corriente"),
        ("N° Cuenta:",    "96353421",        "Correo:",      "aibaceta@gruporamaq.cl"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(bank_rows_data):
        bg = GRIS if i % 2 == 0 else white
        t = Table(
            [[Paragraph(f"<b>{l1}</b>", ps(f"bl1{i}", size=8, bold=True, color=white)),
              Paragraph(v1, ps(f"bv1{i}", size=8)),
              Paragraph(f"<b>{l2}</b>", ps(f"bl2{i}", size=8, bold=True, color=white)),
              Paragraph(v2, ps(f"bv2{i}", size=8))]],
            colWidths=[bank_lw, bank_vw, bank_lw, bank_vw],
        )
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), AZUL_M),
            ("BACKGROUND", (2, 0), (2, 0), AZUL_M),
            ("BACKGROUND", (1, 0), (1, 0), bg),
            ("BACKGROUND", (3, 0), (3, 0), bg),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",(0,0),(-1,-1), 4),
            ("LEFTPADDING",(0, 0), (-1, -1), 6),
            ("RIGHTPADDING",(0,0),(-1,-1), 6),
            ("BOX",        (0, 0), (-1, -1), 0.3, BORDE),
            ("INNERGRID",  (0, 0), (-1, -1), 0.3, BORDE),
        ]))
        story.append(t)

    doc.build(story, canvasmaker=FooterCanvas)



@router.get("/{cotizacion_id}/pdf")
def descargar_pdf(
    cotizacion_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Genera y descarga el PDF de la cotización formal."""
    c = _check_ownership(cotizacion_id, current_user, db)
    config = _get_or_create_config(db)
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.cotizacion_id == cotizacion_id)
        .order_by(ItemCotizacion.item_num)
        .all()
    )
    items_raw = [_item_to_dict(i) for i in items]
    resultado = calcular_cotizacion(items_raw, _config_to_dict(config))

    filename = f"pdf_{cotizacion_id}_{uuid.uuid4().hex[:8]}.pdf"
    filepath = os.path.join(RESULTS_DIR, filename)
    _generar_pdf_formal(c, resultado, filepath, _config_to_dict(config))

    return FileResponse(
        filepath,
        filename=f"CotizacionCliente-{c.numero}.pdf",
        media_type="application/pdf",
    )
