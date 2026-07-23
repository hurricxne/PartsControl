"""
Router de Compras — flujo completo de compra, pre-embarque y embarque.
"""
from datetime import date, datetime, timedelta
from typing import List, Optional

import os
import shutil
import uuid
import json

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy.sql import func

from auth import get_current_user
from empresa_guard import require_empresa
from role_guard import require_rol
from database import get_db
from models.models import (
    Cotizacion, ItemCotizacion, OcCliente, OcProveedor, OcProveedorItem,
    User, Proveedor, Embarque, EmbarqueItem, EmbarqueDocumento, PreEmbarque, PreEmbarqueItem,
    ConfiguracionCotizador, FacturaProveedor, FacturaProveedorItem,
)
from services.pricing_service import (
    calcular_cotizacion, config_efectivo, CLAVES_PRICING, snapshot_desde_config,
)

router = APIRouter(prefix="/compras", tags=["compras"])

DOCS_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "uploads", "docs")
os.makedirs(DOCS_DIR, exist_ok=True)

ALLOWED_EXTENSIONS = {".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx", ".xls", ".xlsx"}
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB


@router.post("/docs/upload")
async def upload_doc(
    file: UploadFile = File(...),
    current_user: "User" = Depends(get_current_user),
):
    """Upload a document file (PDF, image, Office). Returns saved filename."""
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(400, f"Tipo de archivo no permitido: {ext}. Use PDF, imagen o Office.")

    # Read content to check size
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(400, "Archivo demasiado grande (máximo 20 MB)")

    # Save with unique name
    unique_name = f"{uuid.uuid4().hex}{ext}"
    dest = os.path.join(DOCS_DIR, unique_name)
    with open(dest, "wb") as f_out:
        f_out.write(content)

    return {
        "filename": unique_name,
        "original": file.filename,
        "size": len(content),
    }


# ── Feriados Chile + Business-day helpers ────────────────────────────────

def _easter_date(year: int) -> date:
    """Algoritmo de Meeus/Jones/Butcher para calcular Pascua."""
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return date(year, month, day)


def _is_chile_holiday(d: date) -> bool:
    """Retorna True si la fecha es feriado en Chile (fijos + Viernes Santo)."""
    fixed = {
        (1, 1),   # Año Nuevo
        (5, 1),   # Día del Trabajo
        (5, 21),  # Glorias Navales
        (6, 20),  # Día de los Pueblos Indígenas (tercer lunes junio — aproximado fijo)
        (6, 29),  # San Pedro y San Pablo
        (7, 16),  # Virgen del Carmen
        (8, 15),  # Asunción de la Virgen
        (9, 18),  # Independencia
        (9, 19),  # Glorias del Ejército
        (10, 12), # Encuentro de Dos Mundos
        (10, 31), # Día de las Iglesias Evangélicas
        (11, 1),  # Todos los Santos
        (12, 8),  # Inmaculada Concepción
        (12, 25), # Navidad
    }
    if (d.month, d.day) in fixed:
        return True
    # Viernes Santo (2 días antes de Pascua)
    easter = _easter_date(d.year)
    good_friday = easter - timedelta(days=2)
    return d == good_friday


def add_business_days(start_date, days: int) -> date:
    """Retorna la fecha que resulta de agregar *days* días hábiles (Chile)."""
    if isinstance(start_date, datetime):
        current = start_date.date()
    else:
        current = start_date
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5 and not _is_chile_holiday(current):
            added += 1
    return current


def business_days_remaining(deadline: date) -> int:
    """Días hábiles (Chile) hasta deadline. Negativo si ya pasó."""
    today = date.today()
    if deadline == today:
        return 0
    step = 1 if deadline > today else -1
    count = 0
    cursor = today
    while cursor != deadline:
        cursor += timedelta(days=step)
        if cursor.weekday() < 5 and not _is_chile_holiday(cursor):
            count += step
    return count


# ── Schemas ────────────────────────────────────────────────────────────────

class PreEmbarqueUpdate(BaseModel):
    fecha_llegada_est: Optional[str] = None
    notas: Optional[str] = None
    doc_adicional: Optional[str] = None


class EmbarqueUpdate(BaseModel):
    estado: Optional[str] = None
    forwarder: Optional[str] = None
    awb: Optional[str] = None
    awb_numero: Optional[str] = None  # dead-code (esta clase la pisa la de más abajo); se agrega por consistencia
    fecha_despacho: Optional[str] = None
    fecha_llegada_est: Optional[str] = None
    notas: Optional[str] = None
    factura_comercial: Optional[str] = None
    packing_list: Optional[str] = None
    certificado_origen: Optional[str] = None


class OcClienteCreate(BaseModel):
    cotizacion_id: int
    numero_oc: Optional[str] = None
    fecha_oc: Optional[str] = None
    cond_pago: Optional[str] = None
    fecha_entrega: Optional[str] = None
    asesor_id: Optional[int] = None


class OcClienteUpdate(BaseModel):
    numero_oc: Optional[str] = None
    fecha_oc: Optional[str] = None
    cond_pago: Optional[str] = None
    fecha_entrega: Optional[str] = None
    asesor_id: Optional[int] = None


class OcProveedorCreate(BaseModel):
    proveedor: str
    numero_oc: Optional[str] = None
    pais: Optional[str] = None
    moneda: Optional[str] = "USD"
    plazo_dias: Optional[int] = None
    notas: Optional[str] = None
    # Origen de la compra: 'internacional' (embarque) o 'nacional' (camión + guía).
    # Gobierna el camino físico y la UI; el default deja el flujo internacional intacto.
    tipo_origen: Optional[str] = "internacional"


class OcProveedorUpdate(BaseModel):
    proveedor: Optional[str] = None
    numero_oc: Optional[str] = None
    pais: Optional[str] = None
    moneda: Optional[str] = None
    estado: Optional[str] = None
    plazo_dias: Optional[int] = None
    awb: Optional[str] = None
    notas: Optional[str] = None


class ItemPlazo(BaseModel):
    id: int
    plazo_dias_prov: Optional[int] = None
    plazo_entrega_max: Optional[int] = None


class AsignacionItems(BaseModel):
    item_ids: List[int]
    oc_cliente_id: int
    item_plazos: Optional[List[ItemPlazo]] = None


class PreparadoItems(BaseModel):
    item_ids: List[int]


class InvoxItem(BaseModel):
    oc_proveedor_id: int
    numero_invox: str


class CerrarPreEmbarqueBody(BaseModel):
    awb: Optional[str] = None
    # N° AWB/BL escrito a mano (≠ awb, que es el filename del adjunto). Acotado a 100
    # para no exceder la columna VARCHAR(100): un texto más largo debe rebotar como
    # 422 de validación y NO como 500 (DataError 1406) al intentar insertarlo.
    awb_numero: Optional[str] = Field(default=None, max_length=100)
    forwarder: Optional[str] = None
    fecha_despacho: Optional[str] = None
    fecha_llegada_est: Optional[str] = None
    packing_list: Optional[str] = None
    certificado_origen: Optional[str] = None
    notas: Optional[str] = None
    invox_items: Optional[List[InvoxItem]] = None



class PreparadoParcialItem(BaseModel):
    item_id: int
    cantidad: Optional[float] = None

class PreparadoParcialItems(BaseModel):
    items: List[PreparadoParcialItem]

class PreEmbarqueCreate(BaseModel):
    item_ids: List[int]
    notas: Optional[str] = None



class EmbarqueCreate(BaseModel):
    pre_embarque_id: Optional[int] = None
    item_ids: List[int]
    forwarder: Optional[str] = None
    awb: Optional[str] = None
    fecha_despacho: Optional[str] = None
    fecha_llegada_est: Optional[str] = None
    factura_comercial: Optional[str] = None
    packing_list: Optional[str] = None
    certificado_origen: Optional[str] = None
    doc_adicional: Optional[str] = None
    notas: Optional[str] = None


class EmbarqueUpdate(BaseModel):
    estado: Optional[str] = None
    forwarder: Optional[str] = None
    awb: Optional[str] = None
    # N° AWB/BL escrito a mano (≠ awb, que es el filename del adjunto). Acotado a 100
    # para no exceder la columna VARCHAR(100): un texto más largo debe rebotar como
    # 422 de validación y NO como 500 (DataError 1406) al intentar persistirlo.
    awb_numero: Optional[str] = Field(default=None, max_length=100)
    fecha_despacho: Optional[str] = None
    fecha_llegada_est: Optional[str] = None
    factura_comercial: Optional[str] = None
    packing_list: Optional[str] = None
    certificado_origen: Optional[str] = None
    doc_adicional: Optional[str] = None
    notas: Optional[str] = None


class ProveedorCreate(BaseModel):
    nombre: str
    tipo: Optional[str] = "SWIFT"
    pais: Optional[str] = None
    moneda: Optional[str] = None
    contacto: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    sitio_web: Optional[str] = None
    notas: Optional[str] = None


class ProveedorUpdate(BaseModel):
    nombre: Optional[str] = None
    tipo: Optional[str] = None
    pais: Optional[str] = None
    moneda: Optional[str] = None
    contacto: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    sitio_web: Optional[str] = None
    notas: Optional[str] = None


# ── Numero generators ──────────────────────────────────────────────────────

def _next_ocp_numero(db: Session) -> str:
    year = date.today().year
    last = (
        db.query(OcProveedor)
        .filter(OcProveedor.numero.like(f"OCP-{year}-%"))
        .order_by(OcProveedor.id.desc())
        .first()
    )
    seq = int(last.numero.split("-")[-1]) + 1 if last else 1
    return f"OCP-{year}-{seq:04d}"


def _next_pre_numero(db: Session) -> str:
    year = date.today().year
    last = (
        db.query(PreEmbarque)
        .filter(PreEmbarque.numero.like(f"PRE-{year}-%"))
        .order_by(PreEmbarque.id.desc())
        .first()
    )
    seq = int(last.numero.split("-")[-1]) + 1 if last else 1
    return f"PRE-{year}-{seq:04d}"


def _next_emb_numero(db: Session) -> str:
    year = date.today().year
    last = (
        db.query(Embarque)
        .filter(Embarque.numero.like(f"EMB-{year}-%"))
        .order_by(Embarque.id.desc())
        .first()
    )
    seq = int(last.numero.split("-")[-1]) + 1 if last else 1
    return f"EMB-{year}-{seq:04d}"


# ── Pricing helper ─────────────────────────────────────────────────────────

def _cfg_to_dict(cfg) -> dict:
    if cfg is None:
        return {}
    return {
        "tipo_cambio_usd": cfg.tipo_cambio_usd,
        "costo_shipping_usd_kg": cfg.costo_shipping_usd_kg,
        "adicionales_shipping_usd": cfg.adicionales_shipping_usd,
        "costo_agencia_pct": cfg.costo_agencia_pct,
        "costo_agencia_minimo_clp": cfg.costo_agencia_minimo_clp,
        "desconsolidado_clp": cfg.desconsolidado_clp,
        "bodegaje_clp": cfg.bodegaje_clp,
        "margen_venta_pct": cfg.margen_venta_pct,
    }


def _item_obj_to_dict(item: ItemCotizacion) -> dict:
    """Convert ORM item to the flat dict expected by calcular_cotizacion."""
    return {
        "id": item.id,
        "cantidad": item.cantidad or 0,
        "precio_unit_cotizacion": item.precio_unit_cotizacion or 0,
        "peso_unit_lbs": item.peso_unit_lbs or 0,
        "margen_pct": item.margen_pct,
    }


def _calc_pricing_for_items(items: list, cfg, db: Session) -> dict:
    """
    Run calcular_cotizacion per cotizacion group.
    Returns mapping item_id -> total_venta_clp.
    """
    from collections import defaultdict

    # group by cotizacion_id
    groups = defaultdict(list)
    for item in items:
        groups[item.cotizacion_id].append(item)

    result = {}
    cfg_dict = _cfg_to_dict(cfg)
    for cot_id, group_items in groups.items():
        item_dicts = [_item_obj_to_dict(i) for i in group_items]
        _cot = group_items[0].cotizacion if group_items else None
        _cfg = config_efectivo(getattr(_cot, "pricing_snapshot", None), cfg_dict)
        calc = calcular_cotizacion(item_dicts, {**_cfg, "origen": (_cot.origen if _cot else None) or "costo"})
        for calc_item in calc.get("items", []):
            result[calc_item["id"]] = calc_item.get("total_venta_clp", 0)
    return result


def _item_to_dict(item: ItemCotizacion, ocp_id: Optional[int],
                  asig, cfg, total_venta_clp: float, fecha_ref) -> dict:
    """
    Build full item dict for API responses.
    asig: OcProveedorItem or None
    fecha_ref: fallback datetime when asig.fecha_asignacion is None
    """
    plazo_max = (
        item.plazo_entrega_max
        if item.plazo_entrega_max is not None
        else (cfg.plazo_max_default if cfg else 15)
    )

    # Compute deadline
    asig_date = None
    if asig is not None:
        asig_date = asig.fecha_asignacion
    if asig_date is None:
        asig_date = fecha_ref

    dias_restantes = None
    if asig_date is not None and plazo_max:
        deadline = add_business_days(asig_date, plazo_max)
        dias_restantes = business_days_remaining(deadline)

    ocp_numero = None
    ocp_numero_oc = None
    ocp_proveedor = None
    ocp_pais = None
    if ocp_id:
        # will be populated by caller if needed
        pass

    return {
        "id": item.id,
        "cotizacion_id": item.cotizacion_id,
        "item_num": item.item_num,
        "numero_parte": item.numero_parte or "",
        "descripcion": item.descripcion or item.nombre_cat or "",
        "marca": item.marca or "",
        "cantidad": item.cantidad or 1,
        "precio_unit_cotizacion": item.precio_unit_cotizacion or 0,
        "total_cotizacion": item.total_cotizacion or 0,
        "peso_unit_lbs": item.peso_unit_lbs or 0,
        "plazo_entrega_min": item.plazo_entrega_min,
        "plazo_entrega_max": plazo_max,
        "estado_item": item.estado_item or "ingresado",
        "total_venta_clp": total_venta_clp,
        "dias_restantes": dias_restantes,
        "oc_proveedor_id": ocp_id,
        "ocp_numero": ocp_numero,
        "ocp_numero_oc": ocp_numero_oc,
        "ocp_proveedor": ocp_proveedor,
        "ocp_pais": ocp_pais,
        "fecha_asignacion": asig_date.isoformat() if asig_date else None,
        "plazo_dias_prov": asig.plazo_dias_prov if asig else None,
        "ocp_item_id": asig.id if asig else None,
        "unit_price_usd": None,  # populated separately when needed
        "dias_transcurridos": (datetime.utcnow().date() - asig_date.date()).days if asig_date else None,
    }


def _enrich_ocp(d: dict, ocp) -> dict:
    """Add all OCP fields (both ocp_* and oc_proveedor_* aliases) to an item dict."""
    if ocp:
        d["ocp_numero"]         = ocp.numero or ""
        d["ocp_numero_oc"]      = ocp.numero_oc or ""
        d["ocp_proveedor"]      = ocp.proveedor or ""
        d["ocp_pais"]           = ocp.pais or ""
        d["ocp_moneda"]         = ocp.moneda or "USD"
        # frontend aliases used by SeguimientoPage / PreEmbarquesPage / EmbarquesPage
        d["oc_proveedor_numero"] = ocp.numero or ""
        d["oc_proveedor_nombre"] = ocp.proveedor or ""
        d["oc_proveedor_pais"]   = ocp.pais or ""
        d["numero_oc_prov"]      = ocp.numero_oc or ""
    else:
        d["ocp_numero"]          = ""
        d["ocp_numero_oc"]       = ""
        d["ocp_proveedor"]       = ""
        d["ocp_pais"]            = ""
        d["ocp_moneda"]          = "USD"
        d["oc_proveedor_numero"] = ""
        d["oc_proveedor_nombre"] = ""
        d["oc_proveedor_pais"]   = ""
        d["numero_oc_prov"]      = ""
    return d


# ── Endpoints ──────────────────────────────────────────────────────────────

@router.get("/counts")
def get_counts(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # compras = cerrados sin asignar a OC-Proveedor
    cerrados = db.query(ItemCotizacion).filter(ItemCotizacion.estado_item == "cerrado").all()
    sin_ocp = sum(
        1 for i in cerrados
        if not db.query(OcProveedorItem).filter(OcProveedorItem.item_cotizacion_id == i.id).first()
    )
    seguimiento = db.query(ItemCotizacion).filter(ItemCotizacion.estado_item == "comprado").count()
    preparados = db.query(ItemCotizacion).filter(ItemCotizacion.estado_item == "preparado").count()
    # Embarques reales: abierto = armando/despachado/etc; cerrado = entregado
    estados_abierto = ["armando", "despachado", "en_aduana", "en_bodega", "en_transito"]
    embarques_abiertos = db.query(Embarque).filter(Embarque.estado.in_(estados_abierto)).count()
    embarques_cerrados = db.query(Embarque).filter(Embarque.estado == "entregado").count()
    return {
        "compras": sin_ocp,
        "seguimiento": seguimiento,
        "embarques": preparados,
        "embarques_abiertos": embarques_abiertos,
        "embarques_cerrados": embarques_cerrados,
    }


@router.post("/oc-cliente", status_code=201)
def crear_oc_cliente(
    body: OcClienteCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cot = db.query(Cotizacion).filter(Cotizacion.id == body.cotizacion_id).first()
    if not cot:
        raise HTTPException(404, "Cotizacion no encontrada")
    if not (body.numero_oc or "").strip():
        raise HTTPException(400, "El N° OC del cliente es obligatorio")
    # Idempotente por cotización: el cierre de venta hace 2 pasos (crear OC y luego
    # avanzar la fase); si el 2° falló y el usuario reintenta, se devuelve la OC ya
    # creada en vez de duplicarla (los datos se corrigen después con "Editar OC").
    oc_existente = (
        db.query(OcCliente)
        .filter(OcCliente.cotizacion_id == body.cotizacion_id)
        .first()
    )
    if oc_existente:
        return {"id": oc_existente.id, "cotizacion_id": oc_existente.cotizacion_id}
    oc = OcCliente(**body.dict())
    db.add(oc)
    # Congelar la "foto" de precios en el momento del Cierre de Venta: desde aquí el total
    # de la cotización NO se mueve aunque después cambien el dólar u otros parámetros
    # globales del cotizador. Solo se congela al crear la OC (venta que se cierra a partir
    # de esta función); las ventas cerradas ANTES quedan con snapshot NULL y siguen usando
    # el config global (decisión del dueño: congelar solo de aquí en adelante).
    if not cot.pricing_snapshot:
        cfg_obj = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
        if cfg_obj is not None:
            snap = snapshot_desde_config({k: getattr(cfg_obj, k, None) for k in CLAVES_PRICING})
            cot.pricing_snapshot = json.dumps(snap)
            cot.pricing_snapshot_at = func.now()
    # Mark all items of this cotizacion as cerrado
    db.query(ItemCotizacion).filter(
        ItemCotizacion.cotizacion_id == body.cotizacion_id,
        ItemCotizacion.estado_item == "ingresado",
    ).update({"estado_item": "cerrado"})
    db.commit()
    db.refresh(oc)

    # C.4 — Notify abastecimiento of new OC-Cliente
    try:
        from notificaciones import crear_notificacion
        n_items = db.query(ItemCotizacion).filter(
            ItemCotizacion.cotizacion_id == body.cotizacion_id,
            ItemCotizacion.estado_item == "cerrado",
        ).count()
        fecha_str = str(oc.fecha_entrega) if oc.fecha_entrega else "sin definir"
        crear_notificacion(
            db=db,
            rol="abastecimiento",
            severidad="info",
            titulo=f"Nueva OC Cliente N°{oc.numero_oc or oc.id}",
            mensaje=f"{n_items} ítems pendientes de asignar a proveedor — plazo cliente {fecha_str}",
            entidad_tipo="oc_cliente",
            entidad_id=oc.id,
            link=f"/compras?oc_cliente={oc.id}",
            regla=f"nueva_oc_cliente_{oc.id}",
        )
    except Exception as e:
        print(f"[warn] notificacion oc-cliente: {e}")

    return {"id": oc.id, "cotizacion_id": oc.cotizacion_id}


@router.put(
    "/oc-cliente/{oc_id}",
    dependencies=[
        Depends(require_empresa("mineria")),
        Depends(require_rol("comercial", "contabilidad", "admin")),
    ],
)
def actualizar_oc_cliente(
    oc_id: int,
    body: OcClienteUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Edita ex-post los datos de la OC del cliente (N° OC, fecha, condiciones, entrega, asesor).

    Restringido a Grupo AM (empresa 'mineria') y a los roles comercial/contabilidad/admin
    (ver role_guard.py: el candado de rol es permisivo hasta que exista User.rol). Se expone en las
    páginas Ventas y Ventas — Contabilidad para corregir una OC que se cerró sin todos sus datos.
    """
    oc = db.query(OcCliente).filter(OcCliente.id == oc_id).first()
    if not oc:
        raise HTTPException(404, "OC-Cliente no encontrada")
    # Valores que la guía SII referencia (tipo 801): se comparan tras aplicar el
    # body para saber si el PUT intenta cambiarlos (ver guard más abajo).
    numero_oc_original, fecha_oc_original = oc.numero_oc, oc.fecha_oc
    if body.numero_oc is not None:
        if not body.numero_oc.strip():
            raise HTTPException(400, "El N° OC del cliente es obligatorio")
        oc.numero_oc = body.numero_oc.strip()
    if body.fecha_oc is not None:
        oc.fecha_oc = body.fecha_oc or None
    if body.cond_pago is not None:
        oc.cond_pago = body.cond_pago or None
    if body.fecha_entrega is not None:
        oc.fecha_entrega = body.fecha_entrega or None
    # asesor_id distingue "no enviado" (no tocar) de "null explícito" (desasignar):
    # el modal manda null cuando el usuario elige "— sin asesor —", y con el patrón
    # if-not-None a secas esa desasignación se tragaba en silencio con un 200.
    if "asesor_id" in body.model_fields_set:
        if body.asesor_id is None:
            oc.asesor_id = None
        else:
            # Mismo criterio que GET /auth/users (que alimenta el selector): el asesor
            # debe existir, estar activo y ser de la misma empresa — un id de otra
            # empresa se colaría al nombre del asesor que muestran Ventas/Contabilidad.
            asesor = (
                db.query(User)
                .filter(
                    User.id == body.asesor_id,
                    User.is_active == 1,
                    User.empresa == (current_user.empresa or "mineria"),
                )
                .first()
            )
            if not asesor:
                raise HTTPException(400, "Asesor inválido: debe ser un usuario activo de tu empresa")
            oc.asesor_id = body.asesor_id
    # La guía de despacho electrónica (SII 52) referencia esta OC (tipo 801, N° +
    # fecha): con una guía viva emitida (o en emisión), cambiar el N°/fecha de la
    # OC dejaría el documento legal desincronizado del sistema. Se bloquea igual
    # que el folio en despachos (la anulación del DTE se gestiona en Wasabil).
    if (oc.numero_oc, oc.fecha_oc) != (numero_oc_original, fecha_oc_original):
        from wasabil_dte.models import WasabilDte, STATUS_EMITIDO, STATUS_PROCESANDO, STATUS_PENDIENTE
        from wasabil_dte.service import claim_vigente
        from models.models import Despacho
        dtes = (
            db.query(WasabilDte)
            .join(Despacho, Despacho.id == WasabilDte.despacho_id)
            .filter(Despacho.oc_cliente_id == oc.id)
            .all()
        )
        for dte in dtes:
            viva = (
                dte.status_id == STATUS_EMITIDO
                or claim_vigente(dte)
                or (dte.uuid and dte.status_id in (STATUS_PROCESANDO, STATUS_PENDIENTE))
            )
            if viva:
                raise HTTPException(
                    409,
                    "Esta OC ya tiene guía de despacho electrónica emitida"
                    + (f" (folio {dte.folio})" if dte.folio else " (en emisión)")
                    + ": el N° y la fecha de la OC quedaron referenciados ante el SII y no se editan. "
                    "Los demás campos (condiciones, entrega, asesor) sí se pueden corregir.",
                )
    db.commit()
    return {"ok": True}


@router.get("/oc-cliente")
def listar_oc_clientes(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cfg = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
    oc_list = db.query(OcCliente).order_by(OcCliente.created_at.desc()).all()
    result = []

    for oc in oc_list:
        cot = db.query(Cotizacion).filter(Cotizacion.id == oc.cotizacion_id).first()
        items_db = (
            db.query(ItemCotizacion)
            .filter(
                ItemCotizacion.cotizacion_id == oc.cotizacion_id,
                ItemCotizacion.estado_item.in_(
                    ["cerrado", "comprado", "preparado", "pre_embarcado", "embarcado"]
                ),
            )
            .all()
        )

        # Build pricing map for this cotizacion
        pricing_map = _calc_pricing_for_items(items_db, cfg, db)

        items_out = []
        nombre_oc_prov_set = set()
        for item in items_db:
            asig = (
                db.query(OcProveedorItem)
                .filter(
                    OcProveedorItem.item_cotizacion_id == item.id,
                    OcProveedorItem.oc_cliente_id == oc.id,
                )
                .first()
            )
            ocp_id = asig.oc_proveedor_id if asig else None
            ocp = None
            if ocp_id:
                ocp = db.query(OcProveedor).filter(OcProveedor.id == ocp_id).first()
                if ocp:
                    nombre_oc_prov_set.add(ocp.numero or "")

            total_venta = pricing_map.get(item.id, 0)
            d = _item_to_dict(item, ocp_id, asig, cfg, total_venta, oc.created_at)
            _enrich_ocp(d, ocp)
            items_out.append(d)

        # totales for this OC
        total_neto = sum(i.get("total_venta_clp", 0) for i in items_out)

        result.append({
            "id": oc.id,
            "cotizacion_id": oc.cotizacion_id,
            "numero_cot": cot.numero if cot else "",
            "cliente": cot.cliente if cot else "",
            "rut_cliente": cot.rut_cliente if cot else "",
            "numero_oc": oc.numero_oc or "",
            "fecha_oc": oc.fecha_oc or "",
            "cond_pago": oc.cond_pago or "",
            "fecha_entrega": oc.fecha_entrega or "",
            "total_items": len(items_out),
            "items_con_oc": sum(1 for i in items_out if i.get("oc_proveedor_id")),
            "total_neto_clp": total_neto,
            "nombre_oc_prov": ", ".join(sorted(nombre_oc_prov_set)),
            "items": items_out,
            "created_at": oc.created_at.isoformat() if oc.created_at else None,
        })

    return result


@router.get("/oc-cliente/{oc_id}/excel")
def descargar_oc_cliente_excel(
    oc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Genera y descarga Excel de una OC-Cliente con columnas: Cantidad | N° Parte | Marca."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import FileResponse
    import tempfile

    oc = db.query(OcCliente).filter(OcCliente.id == oc_id).first()
    if not oc:
        raise HTTPException(status_code=404, detail="OC no encontrada")

    cot = db.query(Cotizacion).filter(Cotizacion.id == oc.cotizacion_id).first()

    items = (
        db.query(ItemCotizacion)
        .filter(
            ItemCotizacion.cotizacion_id == oc.cotizacion_id,
            ItemCotizacion.estado_item.in_(
                ["cerrado", "comprado", "preparado", "pre_embarcado", "embarcado"]
            ),
        )
        .order_by(ItemCotizacion.item_num)
        .all()
    )

    # ── Crear workbook ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "OC"
    ws.sheet_view.showGridLines = False

    ORNG  = "E87120"
    WHITE = "FFFFFF"
    LGRAY = "F2F2F2"

    def hdr(cell, text):
        cell.value = text
        cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=ORNG)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def cell_val(cell, val, align="center", bg=None):
        cell.value = val
        cell.font = Font(size=9, name="Calibri")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)

    # Anchos: A=12 (Cantidad), B=25 (N° Parte), C=20 (Marca)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 22

    # ── Fila 1: Título ──
    ws.row_dimensions[1].height = 18
    ws.merge_cells("A1:C1")
    numero = oc.numero_oc or (cot.numero if cot else str(oc_id))
    cliente = cot.cliente if cot else ""
    ws["A1"].value = f"OC-{numero}  |  {cliente}"
    ws["A1"].font = Font(bold=True, color=ORNG, size=12, name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    # ── Fila 2: Headers ──
    ws.row_dimensions[2].height = 16
    hdr(ws["A2"], "Cantidad")
    hdr(ws["B2"], "N° Parte")
    hdr(ws["C2"], "Marca")

    # ── Filas de items ──
    for i, item in enumerate(items):
        row = 3 + i
        ws.row_dimensions[row].height = 14
        bg = LGRAY if i % 2 == 0 else None

        cell_val(ws[f"A{row}"], item.cantidad or 1, align="center", bg=bg)
        cell_val(ws[f"B{row}"], item.numero_parte or "", align="left", bg=bg)
        ws[f"B{row}"].font = Font(bold=True, size=9, name="Calibri")
        cell_val(ws[f"C{row}"], item.marca or "", align="left", bg=bg)

    # ── Guardar en temp ──
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.close()

    filename = f"OC-{numero}.xlsx"
    return FileResponse(
        tmp.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@router.post("/oc-proveedor/{ocp_id}/items")
def asignar_items(
    ocp_id: int,
    body: AsignacionItems,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oc = db.query(OcProveedor).filter(OcProveedor.id == ocp_id).first()
    if not oc:
        raise HTTPException(404, "OC-Proveedor no encontrada")

    # Build plazo map
    plazo_map = {}
    plazo_entrega_map = {}
    if body.item_plazos:
        for ip in body.item_plazos:
            if ip.plazo_dias_prov is not None:
                plazo_map[ip.id] = ip.plazo_dias_prov
            if ip.plazo_entrega_max is not None:
                plazo_entrega_map[ip.id] = ip.plazo_entrega_max

    now = datetime.utcnow()
    added = 0
    for item_id in body.item_ids:
        existing = (
            db.query(OcProveedorItem)
            .filter(
                OcProveedorItem.item_cotizacion_id == item_id,
                OcProveedorItem.oc_cliente_id == body.oc_cliente_id,
            )
            .first()
        )
        if existing:
            if existing.oc_proveedor_id != ocp_id:
                existing.oc_proveedor_id = ocp_id
                existing.fecha_asignacion = now
                added += 1
            # Update plazos if provided
            if item_id in plazo_map:
                existing.plazo_dias_prov = plazo_map[item_id]
        else:
            asig = OcProveedorItem(
                oc_proveedor_id=ocp_id,
                oc_cliente_id=body.oc_cliente_id,
                item_cotizacion_id=item_id,
                fecha_asignacion=now,
                plazo_dias_prov=plazo_map.get(item_id),
            )
            db.add(asig)
            added += 1

        # Update item plazo_entrega_max if provided
        if item_id in plazo_entrega_map:
            item_obj = db.query(ItemCotizacion).filter(ItemCotizacion.id == item_id).first()
            if item_obj:
                item_obj.plazo_entrega_max = plazo_entrega_map[item_id]

        # Update estado_item
        item_obj = db.query(ItemCotizacion).filter(ItemCotizacion.id == item_id).first()
        if item_obj:
            item_obj.estado_item = "comprado"

    db.commit()
    return {"ok": True, "asignados": added}


@router.get("/items/comprados")
def get_items_comprados(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cfg = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.estado_item == "comprado")
        .all()
    )
    pricing_map = _calc_pricing_for_items(items, cfg, db)

    # Group by oc_proveedor
    from collections import defaultdict
    groups = defaultdict(list)
    ocp_cache = {}

    for item in items:
        asig = (
            db.query(OcProveedorItem)
            .filter(OcProveedorItem.item_cotizacion_id == item.id)
            .first()
        )
        ocp_id = asig.oc_proveedor_id if asig else None
        groups[ocp_id].append((item, asig))
        if ocp_id and ocp_id not in ocp_cache:
            ocp_cache[ocp_id] = db.query(OcProveedor).filter(OcProveedor.id == ocp_id).first()

    result = []
    for ocp_id, group in groups.items():
        ocp = ocp_cache.get(ocp_id)
        item_list = []
        for item, asig in group:
            total_venta = pricing_map.get(item.id, 0)
            d = _item_to_dict(item, ocp_id, asig, cfg, total_venta,
                              asig.created_at if asig else None)
            _enrich_ocp(d, ocp)
            # enrich with cotizacion + OC-Cliente data
            cot = db.query(Cotizacion).filter(Cotizacion.id == item.cotizacion_id).first()
            d["numero_cot"] = cot.numero if cot else ""
            d["cliente"] = cot.cliente if cot else ""
            occ = db.query(OcCliente).filter(OcCliente.id == asig.oc_cliente_id).first() if asig and asig.oc_cliente_id else None
            d["numero_oc_cliente"] = occ.numero_oc if occ else ""
            item_list.append(d)
        result.append({
            "oc_proveedor_id": ocp_id,
            "numero": ocp.numero if ocp else None,
            "oc_proveedor_numero": ocp.numero if ocp else "",
            "numero_oc_prov": ocp.numero_oc if ocp else None,
            "proveedor": ocp.proveedor if ocp else "Sin asignar",
            "oc_proveedor_nombre": ocp.proveedor if ocp else "Sin asignar",
            "pais": ocp.pais if ocp else "",
            "moneda": ocp.moneda if ocp else "USD",
            # Origen de la OC: los nacionales muestran "Registrar entrega nacional" en
            # Seguimiento (saltan preparado/embarque). Sin ocp → internacional histórico.
            "tipo_origen": ocp.tipo_origen if ocp else "internacional",
            "items": item_list,
        })

    return result


def _rechazar_items_nacionales(db: Session, item_ids: list) -> None:
    """Los ítems asignados a una OC-Proveedor NACIONAL no pasan por
    preparado/pre-embarque/embarque: su camino físico es 'Registrar entrega
    nacional' en Seguimiento. El backend es la autoridad — la UI oculta los
    botones, pero una selección mixta (o una llamada directa al API) no debe
    poder colarlos al pipeline de embarque (hallazgo del dueño probando en vivo)."""
    if not item_ids:
        return
    nacionales = (db.query(ItemCotizacion.numero_parte)
                  .join(OcProveedorItem,
                        OcProveedorItem.item_cotizacion_id == ItemCotizacion.id)
                  .join(OcProveedor,
                        OcProveedor.id == OcProveedorItem.oc_proveedor_id)
                  .filter(ItemCotizacion.id.in_(item_ids),
                          OcProveedor.tipo_origen == "nacional")
                  .all())
    if nacionales:
        partes = ", ".join(sorted({p or "?" for (p,) in nacionales}))
        raise HTTPException(
            400,
            f"Ítem(s) de compra NACIONAL no pasan por embarque: {partes}. "
            "Regístrelos con 'Registrar entrega nacional' en Seguimiento.")


@router.post("/items/preparar")
def preparar_items(
    body: PreparadoItems,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    _rechazar_items_nacionales(db, list(body.item_ids or []))
    updated = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.id.in_(body.item_ids))
        .update({"estado_item": "preparado"}, synchronize_session=False)
    )
    db.commit()
    return {"ok": True, "updated": updated}



@router.post("/items/preparar-parcial")
def preparar_items_parcial(
    body: PreparadoParcialItems,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Marca ítems como preparado, opcionalmente con cantidad parcial.
    Si la cantidad es menor a la del item, divide el item:
    - Original: cantidad reducida → estado 'preparado'
    - Clone: qty restante → estado 'comprado' (queda pendiente)
    """
    _rechazar_items_nacionales(db, [i.item_id for i in body.items])
    processed = 0
    for item_data in body.items:
        item = db.query(ItemCotizacion).filter(
            ItemCotizacion.id == item_data.item_id,
            ItemCotizacion.estado_item == 'comprado',
        ).first()
        if not item:
            continue

        qty = float(item_data.cantidad) if item_data.cantidad else (item.cantidad or 1)
        qty = max(1.0, min(qty, item.cantidad or 1))

        if qty < (item.cantidad or 1):
            remainder = (item.cantidad or 1) - qty
            _clone_item_cotizacion(item, remainder, 'comprado', db)
            item.cantidad = qty
            item.total_cotizacion = qty * (item.precio_unit_cotizacion or 0)

        item.estado_item = 'preparado'
        processed += 1

    db.commit()
    return {"ok": True, "processed": processed}

@router.get("/items/preparados")
def get_items_preparados(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cfg = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
    items = (
        db.query(ItemCotizacion)
        .filter(ItemCotizacion.estado_item == "preparado")
        .all()
    )
    pricing_map = _calc_pricing_for_items(items, cfg, db)

    result = []
    for item in items:
        asig = (
            db.query(OcProveedorItem)
            .filter(OcProveedorItem.item_cotizacion_id == item.id)
            .first()
        )
        ocp_id = asig.oc_proveedor_id if asig else None
        ocp = None
        if ocp_id:
            ocp = db.query(OcProveedor).filter(OcProveedor.id == ocp_id).first()

        total_venta = pricing_map.get(item.id, 0)
        d = _item_to_dict(item, ocp_id, asig, cfg, total_venta,
                          asig.created_at if asig else None)
        _enrich_ocp(d, ocp)
        result.append(d)

    return result


def _clone_item_cotizacion(item, qty: float, estado: str, db):
    """Clone an ItemCotizacion with a new quantity and estado. (helper interno)"""
    from models.models import ItemCotizacion as _IC
    new_item = _IC(
        cotizacion_id         = item.cotizacion_id,
        item_num              = item.item_num,
        descripcion           = item.descripcion,
        numero_parte          = item.numero_parte,
        marca                 = item.marca,
        cantidad              = qty,
        precio_unit_cotizacion = item.precio_unit_cotizacion,
        total_cotizacion      = qty * (item.precio_unit_cotizacion or 0),
        plazo                 = item.plazo,
        peso_unit_lbs         = item.peso_unit_lbs,
        margen_pct            = item.margen_pct,
        precio_finning        = item.precio_finning,
        plazo_entrega_min     = item.plazo_entrega_min,
        plazo_entrega_max     = item.plazo_entrega_max,
        estado_item           = estado,
        nombre_cat            = item.nombre_cat,
        precio_cat            = item.precio_cat,
        moneda_cat            = item.moneda_cat,
        retiro_estimado       = item.retiro_estimado,
        url_cat               = item.url_cat,
        imagen_url            = item.imagen_url,
        encontrado            = item.encontrado,
    )
    db.add(new_item)
    db.flush()

    # Clone OcProveedorItem
    asig = db.query(OcProveedorItem).filter_by(item_cotizacion_id=item.id).first()
    if asig:
        new_asig = OcProveedorItem(
            oc_proveedor_id  = asig.oc_proveedor_id,
            oc_cliente_id    = asig.oc_cliente_id,
            item_cotizacion_id = new_item.id,
            plazo_dias_prov  = asig.plazo_dias_prov,
            fecha_asignacion = asig.fecha_asignacion,
        )
        db.add(new_asig)
        db.flush()

    return new_item


@router.post("/pre-embarques", status_code=201)
def crear_pre_embarque(
    body: PreEmbarqueCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Cuarto camino de entrada al pipeline: mismo guard que preparar/agregar.
    _rechazar_items_nacionales(db, list(body.item_ids or []))
    numero = _next_pre_numero(db)
    pre = PreEmbarque(numero=numero, notas=body.notas)
    db.add(pre)
    db.flush()

    for item_id in body.item_ids:
        asig = (
            db.query(OcProveedorItem)
            .filter(OcProveedorItem.item_cotizacion_id == item_id)
            .first()
        )
        pi = PreEmbarqueItem(
            pre_embarque_id=pre.id,
            item_cotizacion_id=item_id,
            oc_proveedor_id=asig.oc_proveedor_id if asig else None,
        )
        db.add(pi)

    db.query(ItemCotizacion).filter(
        ItemCotizacion.id.in_(body.item_ids)
    ).update({"estado_item": "pre_embarcado"}, synchronize_session=False)

    db.commit()
    db.refresh(pre)
    return {"id": pre.id, "numero": pre.numero}


@router.put("/pre-embarques/{pre_id}")
def actualizar_pre_embarque(
    pre_id: int,
    body: PreEmbarqueUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Actualiza campos editables de un pre-embarque (fecha_llegada_est, notas)."""
    pre = db.query(PreEmbarque).filter(PreEmbarque.id == pre_id).first()
    if not pre:
        raise HTTPException(status_code=404, detail="Pre-embarque no encontrado")
    if body.fecha_llegada_est is not None:
        pre.fecha_llegada_est = body.fecha_llegada_est or None
    if body.notas is not None:
        pre.notas = body.notas or None
    if body.doc_adicional is not None:
        pre.doc_adicional = body.doc_adicional or None
    db.commit()
    return {"ok": True, "fecha_llegada_est": pre.fecha_llegada_est, "notas": pre.notas}


@router.post("/pre-embarques/{pre_id}/cerrar", status_code=201)
def cerrar_pre_embarque(
    pre_id: int,
    body: CerrarPreEmbarqueBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    pre = db.query(PreEmbarque).filter(PreEmbarque.id == pre_id).first()
    if not pre:
        raise HTTPException(404, "Pre-embarque no encontrado")
    if pre.estado == "embarcado":
        raise HTTPException(400, "Pre-embarque ya fue embarcado")
    # No se puede cerrar un pre-embarque sin items: rompía silenciosamente
    if not pre.items or len(pre.items) == 0:
        raise HTTPException(
            400,
            "No se puede generar un embarque sin items. Agregá items al pre-embarque antes de cerrarlo.",
        )

    # Build factura_comercial: "OCP-2026-001: INV-001, OCP-2026-002: INV-002"
    factura_str = None
    if body.invox_items:
        parts = []
        for ii in body.invox_items:
            ocp = db.query(OcProveedor).filter(OcProveedor.id == ii.oc_proveedor_id).first()
            label = ocp.numero if ocp else f"OCP-{ii.oc_proveedor_id}"
            parts.append(f"{label}: {ii.numero_invox}")
        factura_str = ", ".join(parts)

    numero = _next_emb_numero(db)
    emb = Embarque(
        numero=numero,
        estado="en_bodega_proveedor",  # estado inicial; el usuario lo cambia manualmente
        forwarder=body.forwarder,
        awb=body.awb,
        awb_numero=body.awb_numero,  # N° escribible; independiente del archivo awb
        fecha_despacho=body.fecha_despacho,
        fecha_llegada_est=body.fecha_llegada_est,
        factura_comercial=factura_str,
        packing_list=body.packing_list,
        certificado_origen=body.certificado_origen,
        notas=body.notas,
        pre_embarque_id=pre_id,
    )
    db.add(emb)
    db.flush()

    for pi in pre.items:
        item = db.query(ItemCotizacion).filter(ItemCotizacion.id == pi.item_cotizacion_id).first()
        if not item:
            continue

        qty_despacho = float(pi.cantidad_despacho) if pi.cantidad_despacho else (item.cantidad or 1)
        qty_despacho = max(1.0, min(qty_despacho, item.cantidad or 1))

        if qty_despacho < (item.cantidad or 1):
            remainder = (item.cantidad or 1) - qty_despacho
            _clone_item_cotizacion(item, remainder, 'preparado', db)
            item.cantidad = qty_despacho
            item.total_cotizacion = qty_despacho * (item.precio_unit_cotizacion or 0)

        ei = EmbarqueItem(
            embarque_id=emb.id,
            item_cotizacion_id=item.id,
            oc_proveedor_id=pi.oc_proveedor_id,
        )
        db.add(ei)
        item.estado_item = 'embarcado' 

    pre.estado = "embarcado"
    db.commit()
    db.refresh(emb)
    return {"id": emb.id, "numero": emb.numero}



class AddPreEmbarqueItemBody(BaseModel):
    item_id: int



@router.patch("/pre-embarques/{pre_id}/items/{item_id}")
def actualizar_pre_embarque_item(
    pre_id: int,
    item_id: int,
    body: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Actualiza cantidad_despacho de un ítem en el pre-embarque."""
    pi = db.query(PreEmbarqueItem).filter_by(
        pre_embarque_id=pre_id,
        item_cotizacion_id=item_id,
    ).first()
    if not pi:
        raise HTTPException(404, "Item no encontrado en este pre-embarque")

    if 'cantidad_despacho' in body:
        item = db.query(ItemCotizacion).filter_by(id=item_id).first()
        if item:
            qty = float(body['cantidad_despacho'])
            qty = max(1.0, min(qty, item.cantidad or 1))
            pi.cantidad_despacho = qty
        else:
            pi.cantidad_despacho = body['cantidad_despacho']

    db.commit()
    return {"ok": True, "cantidad_despacho": float(pi.cantidad_despacho) if pi.cantidad_despacho else None}

@router.post("/pre-embarques/{pre_id}/items", status_code=201)
def add_item_to_pre_embarque(
    pre_id: int,
    body: AddPreEmbarqueItemBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Agrega un item preparado a un pre-embarque abierto."""
    pre = db.query(PreEmbarque).filter(PreEmbarque.id == pre_id).first()
    if not pre:
        raise HTTPException(404, "Pre-embarque no encontrado")
    if pre.estado != "en_preparacion":
        raise HTTPException(400, "El pre-embarque ya fue cerrado")

    item = db.query(ItemCotizacion).filter(ItemCotizacion.id == body.item_id).first()
    if not item:
        raise HTTPException(404, "Item no encontrado")
    _rechazar_items_nacionales(db, [item.id])
    if item.estado_item != "preparado":
        raise HTTPException(400, f"El item debe estar en estado 'preparado' (actual: {item.estado_item})")

    # Check not already in this pre-embarque
    existing = db.query(PreEmbarqueItem).filter(
        PreEmbarqueItem.pre_embarque_id == pre_id,
        PreEmbarqueItem.item_cotizacion_id == body.item_id,
    ).first()
    if existing:
        raise HTTPException(400, "El item ya está en este pre-embarque")

    asig = db.query(OcProveedorItem).filter(
        OcProveedorItem.item_cotizacion_id == body.item_id
    ).first()
    pi = PreEmbarqueItem(
        pre_embarque_id=pre_id,
        item_cotizacion_id=body.item_id,
        oc_proveedor_id=asig.oc_proveedor_id if asig else None,
    )
    db.add(pi)
    item.estado_item = "pre_embarcado"
    db.commit()
    return {"ok": True, "pre_embarque_id": pre_id, "item_id": body.item_id}


@router.delete("/pre-embarques/{pre_id}/items/{item_id}", status_code=200)
def remove_item_from_pre_embarque(
    pre_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Quita un item de un pre-embarque (lo devuelve a estado preparado)."""
    pre = db.query(PreEmbarque).filter(PreEmbarque.id == pre_id).first()
    if not pre:
        raise HTTPException(404, "Pre-embarque no encontrado")
    if pre.estado != "en_preparacion":
        raise HTTPException(400, "El pre-embarque ya fue cerrado")

    pi = db.query(PreEmbarqueItem).filter(
        PreEmbarqueItem.pre_embarque_id == pre_id,
        PreEmbarqueItem.item_cotizacion_id == item_id,
    ).first()
    if not pi:
        raise HTTPException(404, "Item no encontrado en este pre-embarque")

    db.delete(pi)
    item = db.query(ItemCotizacion).filter(ItemCotizacion.id == item_id).first()
    if item:
        item.estado_item = "preparado"
    db.commit()
    return {"ok": True}


@router.get("/pre-embarques")
def listar_pre_embarques(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    cfg = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
    pres = db.query(PreEmbarque).order_by(PreEmbarque.id.desc()).all()
    result = []

    for pre in pres:
        items_out = []
        total_usd_exwork = 0.0
        total_clp_exwork = 0.0
        peso_total_kg = 0.0
        dias_list = []

        # Get all items for this pre
        item_objs = []
        for pi in pre.items:
            item = (
                db.query(ItemCotizacion)
                .filter(ItemCotizacion.id == pi.item_cotizacion_id)
                .first()
            )
            if item:
                item_objs.append((pi, item))

        all_items = [i for _, i in item_objs]
        pricing_map = _calc_pricing_for_items(all_items, cfg, db)

        for pi, item in item_objs:
            qty = item.cantidad or 1
            precio_usd = item.precio_unit_cotizacion or 0
            peso_lbs = item.peso_unit_lbs or 0
            peso_kg = peso_lbs * 0.453592 * qty
            total_usd = qty * precio_usd
            total_clp = item.total_cotizacion or 0

            total_usd_exwork += total_usd
            total_clp_exwork += total_clp
            peso_total_kg += peso_kg

            asig = (
                db.query(OcProveedorItem)
                .filter(OcProveedorItem.item_cotizacion_id == item.id)
                .first()
            )
            ocp_id = asig.oc_proveedor_id if asig else pi.oc_proveedor_id
            ocp = None
            if ocp_id:
                ocp = db.query(OcProveedor).filter(OcProveedor.id == ocp_id).first()

            total_venta = pricing_map.get(item.id, 0)
            d = _item_to_dict(item, ocp_id, asig, cfg, total_venta,
                              asig.created_at if asig else pre.created_at)
            d["peso_kg"] = round(peso_kg, 4)
            d["total_usd"] = round(total_usd, 4)
            _enrich_ocp(d, ocp)

            if d.get("dias_restantes") is not None:
                dias_list.append(d["dias_restantes"])

            # A-3.2: populate unit_price_usd from FacturaProveedorItem
            fpi_pre = None
            if asig:
                fpi_pre = db.query(FacturaProveedorItem).filter(
                    FacturaProveedorItem.ocp_item_id == asig.id
                ).first()
            d["unit_price_usd"] = float(fpi_pre.unit_price_usd) if fpi_pre and fpi_pre.unit_price_usd else None

            items_out.append(d)

        # find associated embarque awb
        emb = (
            db.query(Embarque)
            .filter(Embarque.pre_embarque_id == pre.id)
            .first()
        )

        # Build unique proveedores summary
        seen_prov = {}
        for d in items_out:
            ocp_id_d = d.get("oc_proveedor_id")
            if ocp_id_d and ocp_id_d not in seen_prov:
                seen_prov[ocp_id_d] = {
                    "oc_proveedor_id": ocp_id_d,
                    "oc_proveedor_numero": d.get("oc_proveedor_numero") or d.get("ocp_numero") or "",
                    "oc_proveedor_nombre": d.get("oc_proveedor_nombre") or d.get("ocp_proveedor") or "",
                    "oc_proveedor_pais": d.get("oc_proveedor_pais") or d.get("ocp_pais") or "",
                    "item_count": sum(1 for x in items_out if x.get("oc_proveedor_id") == ocp_id_d),
                }
        proveedores_list = list(seen_prov.values())

        result.append({
            "id": pre.id,
            "numero": pre.numero,
            "estado": pre.estado,
            "notas": pre.notas or "",
            "doc_adicional": pre.doc_adicional or "",
            "fecha_llegada_est": pre.fecha_llegada_est or "",
            "created_at": pre.created_at.isoformat() if pre.created_at else None,
            "proveedores": proveedores_list,
            "total_usd_exwork": round(total_usd_exwork, 2),
            "total_clp_exwork": round(total_clp_exwork, 2),
            "peso_total_kg": round(peso_total_kg, 4),
            "max_dias_restantes": max(dias_list) if dias_list else None,
            "embarque_id": emb.id if emb else None,
            "embarque_numero": emb.numero if emb else None,
            "embarque_awb": emb.awb if emb else None,
            "items": items_out,
        })

    return result



@router.get("/embarques-list")
def listar_embarques(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    embarques = db.query(Embarque).order_by(Embarque.id.desc()).all()
    result = []
    for emb in embarques:
        pre = None
        if emb.pre_embarque_id:
            pre = db.query(PreEmbarque).filter(PreEmbarque.id == emb.pre_embarque_id).first()
        result.append({
            "id": emb.id,
            "numero": emb.numero,
            "estado": emb.estado,
            "forwarder": emb.forwarder or "",
            "awb": emb.awb or "",
            "awb_numero": emb.awb_numero or "",
            "fecha_despacho": emb.fecha_despacho or "",
            "fecha_llegada_est": emb.fecha_llegada_est or "",
            "factura_comercial": emb.factura_comercial or "",
            "packing_list": emb.packing_list or "",
            "certificado_origen": emb.certificado_origen or "",
                "doc_adicional": emb.doc_adicional or "",
            "notas": emb.notas or "",
            "pre_embarque_id": emb.pre_embarque_id,
            "pre_embarque_numero": pre.numero if pre else None,
            "total_items": len(emb.items),
            "created_at": emb.created_at.isoformat() if emb.created_at else None,
        })
    return result


@router.get("/embarques-list/{emb_id}")
def get_embarque(
    emb_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    emb = db.query(Embarque).filter(Embarque.id == emb_id).first()
    if not emb:
        raise HTTPException(404, "Embarque no encontrado")

    cfg = db.query(ConfiguracionCotizador).filter(ConfiguracionCotizador.id == 1).first()
    pre = None
    if emb.pre_embarque_id:
        pre = db.query(PreEmbarque).filter(PreEmbarque.id == emb.pre_embarque_id).first()

    # Build pricing for all items in this embarque
    all_items = [ei.item_cotizacion for ei in emb.items if ei.item_cotizacion]
    pricing_map = _calc_pricing_for_items(all_items, cfg, db)

    items_out = []
    for ei in emb.items:
        item = ei.item_cotizacion
        if not item:
            continue
        cot = db.query(Cotizacion).filter(Cotizacion.id == item.cotizacion_id).first()
        asig = (
            db.query(OcProveedorItem)
            .filter(OcProveedorItem.item_cotizacion_id == item.id)
            .first()
        )
        ocp_id = ei.oc_proveedor_id or (asig.oc_proveedor_id if asig else None)
        ocp = None
        if ocp_id:
            ocp = db.query(OcProveedor).filter(OcProveedor.id == ocp_id).first()
        total_venta = pricing_map.get(item.id, 0)
        d = _item_to_dict(item, ocp_id, asig, cfg, total_venta,
                          asig.created_at if asig else emb.created_at)
        _enrich_ocp(d, ocp)
        d["embarque_item_id"] = ei.id
        d["numero_cot"] = cot.numero if cot else ""
        d["cliente"] = cot.cliente if cot else ""

        # A-3.4: populate unit_price_usd from FacturaProveedorItem
        fpi = None
        if asig:
            fpi = db.query(FacturaProveedorItem).filter(
                FacturaProveedorItem.ocp_item_id == asig.id
            ).first()
        if fpi and fpi.unit_price_usd:
            d["unit_price_usd"] = float(fpi.unit_price_usd)
        else:
            d["unit_price_usd"] = None

        # A-3.4: OC Cliente numero
        occ = None
        if asig and asig.oc_cliente_id:
            occ = db.query(OcCliente).filter(OcCliente.id == asig.oc_cliente_id).first()
        d["numero_oc_cliente"] = occ.numero_oc if occ else ""

        # A-3.4: Invoice (factura proveedor)
        # Modelo FacturaProveedor usa ocp_id (FK) e invoice_no
        factura = None
        if ocp_id:
            factura = db.query(FacturaProveedor).filter(
                FacturaProveedor.ocp_id == ocp_id
            ).first()
        d["numero_factura"] = factura.invoice_no if factura else ""

        items_out.append(d)

    return {
        "id": emb.id,
        "numero": emb.numero,
        "estado": emb.estado,
        "forwarder": emb.forwarder or "",
        "awb": emb.awb or "",
        "awb_numero": emb.awb_numero or "",
        "fecha_despacho": emb.fecha_despacho or "",
        "fecha_llegada_est": emb.fecha_llegada_est or "",
        "factura_comercial": emb.factura_comercial or "",
        "packing_list": emb.packing_list or "",
        "certificado_origen": emb.certificado_origen or "",
                "doc_adicional": emb.doc_adicional or "",
        "notas": emb.notas or "",
        "pre_embarque_id": emb.pre_embarque_id,
        "pre_embarque_numero": pre.numero if pre else None,
        "created_at": emb.created_at.isoformat() if emb.created_at else None,
        "items": items_out,
    }


@router.put("/embarques-list/{emb_id}")
def actualizar_embarque(
    emb_id: int,
    body: EmbarqueUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Actualiza metadata y/o documentos de un embarque.

    Reglas de máquina de estados:
    - Orden:  en_bodega_proveedor < en_transito < en_aduana < en_bodega < despachado
    - No se puede retroceder desde 'en_bodega' o 'despachado' (estados terminales/finales)
    - 'despachado' solo se setea automáticamente al cerrar todos los despachos
    - Cuando se cambia a 'en_bodega' desde EmbarquesPage: los items vinculados pasan
      automáticamente a estado_item='en_bodega' para quedar disponibles en Despachos
    """
    emb = db.query(Embarque).filter(Embarque.id == emb_id).first()
    if not emb:
        raise HTTPException(404, "Embarque no encontrado")

    estado_anterior = emb.estado
    data = body.dict(exclude_none=True)
    nuevo_estado = data.get("estado")

    ESTADO_ORDER = {
        "en_bodega_proveedor": 1,
        "en_transito": 2,
        "en_aduana": 3,
        "en_bodega": 4,
        "despachado": 5,
    }

    # Validación de transición de estado
    if nuevo_estado and nuevo_estado != estado_anterior:
        # No permitir retroceso desde estados terminales
        if estado_anterior in ("en_bodega", "despachado"):
            anterior_rank = ESTADO_ORDER.get(estado_anterior, 99)
            nuevo_rank = ESTADO_ORDER.get(nuevo_estado, 0)
            if nuevo_rank < anterior_rank:
                raise HTTPException(
                    400,
                    f"No se puede cambiar el estado del embarque a '{nuevo_estado}'. "
                    f"Una vez en '{estado_anterior}' no se puede volver a un estado anterior.",
                )
        # No permitir setear manualmente "despachado" (solo auto)
        if nuevo_estado == "despachado" and estado_anterior != "despachado":
            raise HTTPException(
                400,
                "El estado 'despachado' se setea automáticamente cuando todos los items "
                "del embarque son despachados. No se puede asignar manualmente.",
            )

    # Aplicar campos
    for field, value in data.items():
        setattr(emb, field, value or None)

    # NOTA: NO se auto-promueven items al pasar a "en_bodega" desde EmbarquesPage.
    # El estado "en_bodega" del Embarque solo indica que llego fisicamente.
    # Para promover los items y que aparezcan en Despachos, Bodega debe abrir
    # y cerrar la recepcion fisica (RecepcionEmbarque) -- ahi recien items
    # pasan a estado_item="en_bodega" (o reclamo_proveedor si hay danos/faltantes).

    db.commit()
    return {"ok": True}


class EmbarqueDocIn(BaseModel):
    nombre: str
    archivo: str


@router.get("/embarques-list/{emb_id}/docs")
def list_embarque_docs(
    emb_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Documentos adicionales del embarque (botón "Otros")."""
    docs = (db.query(EmbarqueDocumento)
            .filter(EmbarqueDocumento.embarque_id == emb_id)
            .order_by(EmbarqueDocumento.id).all())
    return [{"id": d.id, "nombre": d.nombre, "archivo": d.archivo,
             "created_at": d.created_at.isoformat() if d.created_at else None} for d in docs]


@router.post("/embarques-list/{emb_id}/docs", status_code=201)
def add_embarque_doc(
    emb_id: int,
    body: EmbarqueDocIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Adjunta un documento ya subido via POST /compras/docs/upload."""
    if not db.query(Embarque).filter(Embarque.id == emb_id).first():
        raise HTTPException(404, "Embarque no encontrado")
    nombre = (body.nombre or "").strip() or body.archivo
    d = EmbarqueDocumento(embarque_id=emb_id, nombre=nombre[:255],
                          archivo=body.archivo, usuario_id=current_user.id)
    db.add(d)
    db.commit()
    db.refresh(d)
    return {"id": d.id, "nombre": d.nombre, "archivo": d.archivo,
            "created_at": d.created_at.isoformat() if d.created_at else None}


@router.delete("/embarques-list/{emb_id}/docs/{doc_id}", status_code=200)
def delete_embarque_doc(
    emb_id: int,
    doc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    d = (db.query(EmbarqueDocumento)
         .filter(EmbarqueDocumento.id == doc_id,
                 EmbarqueDocumento.embarque_id == emb_id).first())
    if not d:
        raise HTTPException(404, "Documento no encontrado")
    db.delete(d)
    db.commit()
    return {"ok": True}


@router.delete("/embarques-list/{emb_id}/items/{embarque_item_id}", status_code=200)
def desembarcar_item(
    emb_id: int,
    embarque_item_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Quita un ítem del embarque y lo devuelve a estado 'pre_embarcado'."""
    emb = db.query(Embarque).filter(Embarque.id == emb_id).first()
    if not emb:
        raise HTTPException(404, "Embarque no encontrado")

    ei = db.query(EmbarqueItem).filter(
        EmbarqueItem.id == embarque_item_id,
        EmbarqueItem.embarque_id == emb_id,
    ).first()
    if not ei:
        raise HTTPException(404, "Item no encontrado en este embarque")

    # Devolver item a estado pre_embarcado
    item = db.query(ItemCotizacion).filter(ItemCotizacion.id == ei.item_cotizacion_id).first()
    if item:
        item.estado_item = "preparado"

    db.delete(ei)
    db.commit()
    return {"ok": True, "item_id": ei.item_cotizacion_id}


@router.put("/oc-proveedor/{oc_id}")
def actualizar_oc_proveedor(
    oc_id: int,
    body: OcProveedorUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    oc = db.query(OcProveedor).filter(OcProveedor.id == oc_id).first()
    if not oc:
        raise HTTPException(404, "OC-Proveedor no encontrada")
    for k, v in body.dict(exclude_none=True).items():
        setattr(oc, k, v)
    db.commit()
    return {"ok": True}


@router.get("/oc-proveedor")
def listar_oc_proveedores(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    ocs = db.query(OcProveedor).order_by(OcProveedor.id.desc()).all()
    result = []
    for oc in ocs:
        items_out = []
        for asig in oc.items:
            item = asig.item_cotizacion
            if not item:
                continue
            oc_cli = (
                db.query(OcCliente)
                .filter(OcCliente.id == asig.oc_cliente_id)
                .first()
            )
            cot = (
                db.query(Cotizacion)
                .filter(Cotizacion.id == oc_cli.cotizacion_id)
                .first()
                if oc_cli
                else None
            )
            d = {
                "id": item.id,
                "item_num": item.item_num,
                "numero_parte": item.numero_parte or "",
                "descripcion": item.descripcion or "",
                "cantidad": item.cantidad or 1,
                "precio_unit_cotizacion": item.precio_unit_cotizacion or 0,
                "total_cotizacion": item.total_cotizacion or 0,
                "estado_item": item.estado_item or "",
                "numero_cot": cot.numero if cot else "",
                "cliente": cot.cliente if cot else "",
                "numero_oc_cliente": oc_cli.numero_oc if oc_cli else "",
            }
            items_out.append(d)

        result.append({
            "id": oc.id,
            "numero": oc.numero,
            "numero_oc": oc.numero_oc or "",
            "proveedor": oc.proveedor,
            "pais": oc.pais or "",
            "moneda": oc.moneda or "USD",
            "estado": oc.estado or "borrador",
            "plazo_dias": oc.plazo_dias,
            "awb": oc.awb or "",
            "notas": oc.notas or "",
            "total_items": len(items_out),
            "items": items_out,
            "created_at": oc.created_at.isoformat() if oc.created_at else None,
        })
    return result


@router.post("/oc-proveedor", status_code=201)
def crear_oc_proveedor(
    body: OcProveedorCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    numero = _next_ocp_numero(db)
    data = body.dict()
    # La columna es NOT NULL: un cliente que mande tipo_origen=null no debe romper el
    # INSERT (backend autoridad). Solo 'nacional' e 'internacional' son válidos.
    if data.get("tipo_origen") not in ("nacional", "internacional"):
        data["tipo_origen"] = "internacional"
    oc = OcProveedor(numero=numero, **data)
    db.add(oc)
    db.commit()
    db.refresh(oc)
    return {"id": oc.id, "numero": oc.numero}


@router.get("/proveedores")
def listar_proveedores(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    provs = db.query(Proveedor).order_by(Proveedor.nombre).all()
    return [
        {
            "id": p.id,
            "nombre": p.nombre,
            "tipo": p.tipo or "SWIFT",
            "pais": p.pais or "",
            "moneda": p.moneda or "",
            "contacto": p.contacto or "",
            "email": p.email or "",
            "telefono": p.telefono or "",
            "sitio_web": p.sitio_web or "",
            "notas": p.notas or "",
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }
        for p in provs
    ]


@router.post("/proveedores", status_code=201)
def crear_proveedor(
    body: ProveedorCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = Proveedor(**body.dict())
    db.add(p)
    db.commit()
    db.refresh(p)
    return {"id": p.id, "nombre": p.nombre}


@router.put("/proveedores/{p_id}")
def actualizar_proveedor(
    p_id: int,
    body: ProveedorUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = db.query(Proveedor).filter(Proveedor.id == p_id).first()
    if not p:
        raise HTTPException(404, "Proveedor no encontrado")
    for k, v in body.dict(exclude_none=True).items():
        setattr(p, k, v)
    db.commit()
    return {"ok": True}


@router.delete("/proveedores/{p_id}", status_code=204)
def eliminar_proveedor(
    p_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    p = db.query(Proveedor).filter(Proveedor.id == p_id).first()
    if not p:
        raise HTTPException(404, "Proveedor no encontrado")
    db.delete(p)
    db.commit()
    return None


# ─── v2: precio USD editable ──────────────────────────────────────────────

class PrecioUsdBody(BaseModel):
    unit_price_usd: float


@router.patch("/items/{item_id}/precio-usd")
def update_precio_usd(
    item_id: int,
    body: PrecioUsdBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Update or create FacturaProveedorItem.unit_price_usd for a given ItemCotizacion.
    Looks for an existing FacturaProveedorItem via OcProveedorItem → FacturaProveedor chain.
    Creates one if not found (linked to the first FacturaProveedor of the OCP, or standalone).
    """
    item = db.query(ItemCotizacion).filter(ItemCotizacion.id == item_id).first()
    if not item:
        raise HTTPException(404, "Item not found")

    # Try to find via OcProveedorItem
    ocp_item = (
        db.query(OcProveedorItem)
        .filter(OcProveedorItem.item_cotizacion_id == item_id)
        .first()
    )

    fpi = None
    if ocp_item:
        fpi = (
            db.query(FacturaProveedorItem)
            .filter(FacturaProveedorItem.ocp_item_id == ocp_item.id)
            .first()
        )

    if fpi:
        fpi.unit_price_usd = body.unit_price_usd
    else:
        # Create a standalone FacturaProveedorItem
        # Find or create a FacturaProveedor for the OCP
        # NOTA: el modelo usa ocp_id (no oc_proveedor_id) e invoice_no (no numero_factura)
        factura = None
        if ocp_item and ocp_item.oc_proveedor_id:
            factura = (
                db.query(FacturaProveedor)
                .filter(FacturaProveedor.ocp_id == ocp_item.oc_proveedor_id)
                .first()
            )
        if not factura:
            # Create a placeholder factura
            factura = FacturaProveedor(
                ocp_id=ocp_item.oc_proveedor_id if ocp_item else None,
                invoice_no=f"AUTO-{item_id}",
            )
            db.add(factura)
            db.flush()

        fpi = FacturaProveedorItem(
            factura_id=factura.id,
            ocp_item_id=ocp_item.id if ocp_item else None,
            descripcion=item.descripcion or item.numero_parte or "",
            qty_facturada=item.cantidad or 1,
            unit_price_usd=body.unit_price_usd,
        )
        db.add(fpi)

    db.commit()
    db.refresh(fpi)
    total_usd = float(fpi.unit_price_usd or 0) * float(item.cantidad or 1)
    return {
        "item_id": item_id,
        "unit_price_usd": float(fpi.unit_price_usd),
        "total_usd": total_usd,
        "factura_proveedor_item_id": fpi.id,
    }
