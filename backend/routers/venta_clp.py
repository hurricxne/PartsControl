"""
Carga de cotizaciones "con precios de VENTA" en CLP (formato GRUPO AM-CAT, ej. 2536).
A diferencia del Excel de proveedor (costo USD -> pipeline), aquí el precio unitario
YA es precio de venta en CLP (margen incluido). Se carga tal cual, sin scraping ni
conversión. Guarda los días (plazo) por ítem. Marca la cotización con origen='venta_clp'.
"""
import os
import tempfile
import openpyxl
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session

from database import get_db
from auth import get_current_user
from models.models import Cotizacion, ItemCotizacion, User
from routers.cotizaciones import _next_cot_numero

router = APIRouter(prefix="/api/cotizaciones", tags=["cotizaciones-venta-clp"])


def _num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(".", "").replace(",", ".").replace(" ", "")
    try:
        return float(s)
    except Exception:
        return None


def parse_venta_clp(path: str) -> dict:
    """Parsea el formato de venta CLP. Devuelve encabezado + items."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    maxc = ws.max_column or 12
    maxr = ws.max_row or 30

    hdr = {"numero_file": None, "cliente": None, "referencia": None,
           "marca": None, "contacto": None, "total_clp": None, "fecha": None}
    # Encabezado: buscar etiquetas en las primeras 6 filas y tomar la celda siguiente
    for r in range(1, min(maxr, 6) + 1):
        for c in range(1, maxc + 1):
            v = ws.cell(r, c).value
            if v is None:
                continue
            s = str(v).strip().lower().rstrip(":")
            nxt = ws.cell(r, c + 1).value
            if s == "cotizacion" or s == "cotización" or s == "cotización":
                hdr["numero_file"] = nxt
            elif s.startswith("cliente"):
                hdr["cliente"] = nxt
            elif s.startswith("referencia"):
                hdr["referencia"] = nxt
            elif s.startswith("marca"):
                hdr["marca"] = nxt
            elif s.startswith("contacto"):
                hdr["contacto"] = nxt
            elif s.startswith("fecha"):
                hdr["fecha"] = nxt
            elif s.startswith("total"):
                hdr["total_clp"] = _num(nxt)

    # Fila de encabezado de la tabla (contiene PN y Precio Unit ... CLP)
    hrow, colmap = None, {}
    for r in range(1, min(maxr, 15) + 1):
        cmap = {}
        for c in range(1, maxc + 1):
            x = str(ws.cell(r, c).value or "").strip().lower()
            if not x:
                continue
            if x == "#":
                cmap["num"] = c
            elif x == "pn" or "parte" in x or x == "sku":
                cmap["pn"] = c
            elif "descrip" in x:
                cmap["desc"] = c
            elif x == "marca":
                cmap["marca"] = c
            elif "cantidad" in x or x == "cant":
                cmap["cant"] = c
            elif "precio" in x:
                cmap["precio"] = c
            elif x.startswith("total"):
                cmap["total"] = c
            elif "plazo" in x or "entrega" in x or "día" in x or "dia" in x:
                cmap["plazo"] = c
        if "pn" in cmap and "precio" in cmap:
            hrow, colmap = r, cmap
            break
    if hrow is None:
        raise HTTPException(status_code=400, detail="No se encontró la tabla de ítems (columnas PN / Precio Unit. CLP).")

    items = []
    for r in range(hrow + 1, maxr + 1):
        first = ws.cell(r, 1).value
        if first is not None and "total" in str(first).strip().lower():
            break
        pn = ws.cell(r, colmap["pn"]).value
        precio = _num(ws.cell(r, colmap["precio"]).value)
        if (pn is None or str(pn).strip() == "") and precio is None:
            continue
        cant = _num(ws.cell(r, colmap["cant"]).value) if "cant" in colmap else 1.0
        items.append({
            "numero_parte": str(pn).strip() if pn is not None else None,
            "descripcion": str(ws.cell(r, colmap["desc"]).value or "").strip() if "desc" in colmap else "",
            "marca": (str(ws.cell(r, colmap["marca"]).value or "").strip() if "marca" in colmap else "")
                     or (str(hdr["marca"]).strip() if hdr["marca"] else ""),
            "cantidad": cant or 1.0,
            "precio_clp": precio or 0.0,
            "plazo": str(ws.cell(r, colmap["plazo"]).value or "").strip() if "plazo" in colmap else "",
        })
    if not items:
        raise HTTPException(status_code=400, detail="No se encontraron ítems en el archivo.")
    hdr["items"] = items
    return hdr


@router.post("/upload-venta", status_code=201)
async def upload_venta_clp(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Sube un Excel con precios de VENTA en CLP (margen incluido). Sin scraping."""
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1])
    try:
        tmp.write(await file.read())
        tmp.close()
        data = parse_venta_clp(tmp.name)
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass

    numero = _next_cot_numero(db)
    ref = data.get("referencia")
    if data.get("numero_file"):
        ref = f"{data['numero_file']}" + (f" · {ref}" if ref else "")
    cot = Cotizacion(
        numero=numero,
        cliente=str(data.get("cliente") or "").strip() or None,
        contacto_cliente=str(data.get("contacto") or "").strip() or None,
        referencia=str(ref or "").strip() or None,
        archivo_original=file.filename,
        total_items=len(data["items"]),
        items_procesados=len(data["items"]),
        items_encontrados=len(data["items"]),
        estado="COMPLETADO",
        origen="venta_clp",
        user_id=current_user.id,
    )
    db.add(cot)
    db.flush()

    for i, it in enumerate(data["items"], 1):
        precio = it["precio_clp"] or 0.0
        _digs = "".join(ch if ch.isdigit() else " " for ch in (it["plazo"] or "")).split()
        _pmin = int(_digs[0]) if len(_digs) >= 1 else None
        _pmax = int(_digs[1]) if len(_digs) >= 2 else _pmin
        db.add(ItemCotizacion(
            cotizacion_id=cot.id,
            item_num=i,
            numero_parte=it["numero_parte"],
            descripcion=it["descripcion"],
            marca=it["marca"],
            cantidad=it["cantidad"],
            precio_unit_cotizacion=precio,   # precio de VENTA CLP (margen incluido)
            total_cotizacion=precio * (it["cantidad"] or 1),
            plazo=it["plazo"] or None,
            plazo_entrega_min=_pmin,
            plazo_entrega_max=_pmax,
            margen_pct=0.0,                  # el margen ya viene incluido
            estado_item="ingresado",
            encontrado=1,
        ))
    db.commit()
    db.refresh(cot)
    return {"id": cot.id, "numero": cot.numero, "origen": "venta_clp",
            "items": len(data["items"]), "cliente": cot.cliente}
