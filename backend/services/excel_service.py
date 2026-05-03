import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Tuple
import re
import os


def parse_excel_cotizacion(filepath: str) -> Tuple[dict, List[dict]]:
    """
    Lee un Excel de cotización y retorna:
    - header_info: dict con Cotización N°, Cliente, Referencia, Fecha
    - items: lista de dicts con los datos de cada parte

    Soporta múltiples formatos:
    - Formato A: filas de metadata arriba, luego fila de headers con ITEM/N° PARTE
    - Formato B (FE/CAT export): fila 1 directamente es el header (Part Number, Item #, etc.)
    """
    df_raw = pd.read_excel(filepath, header=None, dtype=str)

    header_info = {}
    header_row = None

    # Palabras clave que identifican una fila de encabezados de ítems
    PART_KEYWORDS  = ["PARTE", "PART", "PART NUMBER", "N° PARTE", "NUMERO PARTE"]
    ITEM_KEYWORDS  = ["ITEM", "ÍTEM", "ITEM #", "ITEM NO"]

    def row_has_part_col(row_vals):
        return any(any(kw in v for kw in PART_KEYWORDS) for v in row_vals)

    def row_has_item_col(row_vals):
        return any(any(kw in v for kw in ITEM_KEYWORDS) for v in row_vals)

    # Buscar fila de encabezados (buscar PART y algo de ITEM/QTY/DESC en la misma fila)
    for i, row in df_raw.iterrows():
        row_vals = [str(v).strip().upper() for v in row.values if pd.notna(v)]
        if row_has_part_col(row_vals) and (row_has_item_col(row_vals) or
                any("QTY" in v or "CANT" in v or "DESCRIP" in v for v in row_vals)):
            header_row = i
            break

    # Extraer info del header (primeras 10 filas antes de los ítems)
    scan_limit = min(10, header_row if header_row is not None else 10)
    for i in range(scan_limit):
        row = df_raw.iloc[i]
        row_str = " ".join([str(v) for v in row.values if pd.notna(v)])
        row_up = row_str.upper()
        vals = [str(v).strip() for v in row.values if pd.notna(v) and str(v).strip()]

        if "COTIZACI" in row_up:
            nums = re.findall(r'\d+', row_str)
            if nums:
                header_info["numero"] = nums[0]
        if "CLIENTE" in row_up and "cliente" not in header_info:
            if len(vals) >= 2:
                header_info["cliente"] = vals[1]
        if "RUT" in row_up and "rut_cliente" not in header_info:
            if len(vals) >= 2:
                header_info["rut_cliente"] = vals[1]
        if "CONTACTO" in row_up and "contacto_cliente" not in header_info:
            if len(vals) >= 2:
                header_info["contacto_cliente"] = vals[1]
        if "EMAIL" in row_up and "email_cliente" not in header_info:
            if len(vals) >= 2:
                header_info["email_cliente"] = vals[1]
        if ("TELEFONO" in row_up or "TELÉFONO" in row_up) and "telefono_cliente" not in header_info:
            if len(vals) >= 2:
                header_info["telefono_cliente"] = vals[1]
        if "DIRECCI" in row_up and "direccion_cliente" not in header_info:
            if len(vals) >= 2:
                header_info["direccion_cliente"] = vals[1]
        if "REFERENCIA" in row_up and "referencia" not in header_info:
            if len(vals) >= 2:
                header_info["referencia"] = vals[1]

    if header_row is None:
        raise ValueError(
            "No se encontró la fila de encabezados en el Excel. "
            "Asegúrese de que exista una fila con columnas como 'Part Number' / 'N° PARTE', "
            "'Item #' / 'ITEM', 'Description', 'Qty Required' / 'CANTIDAD'."
        )

    # Leer datos desde la fila de encabezados
    df = pd.read_excel(filepath, header=header_row, dtype=str)
    df.columns = [str(col).strip().upper() for col in df.columns]

    # Mapear columnas flexiblemente (soporta español e inglés)
    col_map = {}
    for col in df.columns:
        cu = col.upper()
        if not col_map.get("numero_parte") and ("PARTE" in cu or "PART NUMBER" in cu or cu == "PART"):
            col_map["numero_parte"] = col
        elif not col_map.get("item") and ("ITEM" in cu or "ÍTEM" in cu):
            col_map["item"] = col
        elif not col_map.get("descripcion") and "DESCRIP" in cu:
            col_map["descripcion"] = col
        elif not col_map.get("marca") and ("MARCA" in cu or "BRAND" in cu):
            col_map["marca"] = col
        elif not col_map.get("cantidad") and ("CANT" in cu or "QTY" in cu or "REQUIRED" in cu):
            col_map["cantidad"] = col
        elif not col_map.get("precio_unit") and (
            ("PRECIO" in cu and "UNIT" in cu) or
            ("UNIT" in cu and "PRICE" in cu) or
            cu in ("UNIT PRICE", "PRECIO UNIT.")
        ):
            col_map["precio_unit"] = col
        elif not col_map.get("total") and "TOTAL" in cu and "PRECIO" not in cu and "SUB" not in cu:
            col_map["total"] = col
        elif not col_map.get("plazo") and "PLAZO" in cu:
            col_map["plazo"] = col
        elif not col_map.get("peso_unit") and (
            "WEIGHT" in cu or
            ("PESO" in cu and "UNIT" in cu) or
            cu in ("UNIT (LBS)", "LBS", "UNIT WEIGHT")
        ):
            col_map["peso_unit"] = col

    def safe_float(v):
        try:
            return float(str(v).replace(",", ".").replace(" ", "")) if pd.notna(v) and str(v).strip() not in ("", "nan", "NAN", "None") else None
        except Exception:
            return None

    items = []
    for _, row in df.iterrows():
        np_col = col_map.get("numero_parte", "")
        np_val = row.get(np_col, "") if np_col else ""
        if pd.isna(np_val) or not str(np_val).strip() or str(np_val).strip().upper() in [
            "NAN", "N° PARTE", "PARTE", "PART NUMBER", "PART", "NONE", ""
        ]:
            continue
        np_clean = str(np_val).strip()
        # Saltar filas de totales/subtotales
        if any(k in np_clean.upper() for k in ["SUBTOTAL", "IVA", "TOTAL", "NETO"]):
            continue
        # Número de parte vacío o sólo espacios
        if not np_clean or np_clean in ("-", "—"):
            continue

        items.append({
            "item_num":               safe_float(row.get(col_map.get("item", ""), None)),
            "descripcion":            str(row.get(col_map.get("descripcion", ""), "") or "").strip(),
            "numero_parte":           np_clean,
            "marca":                  str(row.get(col_map.get("marca", ""), "") or "").strip(),
            "cantidad":               safe_float(row.get(col_map.get("cantidad", ""), None)) or 1,
            "precio_unit_cotizacion": safe_float(row.get(col_map.get("precio_unit", ""), None)),
            "total_cotizacion":       safe_float(row.get(col_map.get("total", ""), None)),
            "plazo":                  str(row.get(col_map.get("plazo", ""), "") or "").strip(),
            "peso_unit_lbs":          safe_float(row.get(col_map.get("peso_unit", ""), None)),
        })

    return header_info, items


def format_part_number(raw: str) -> str:
    """
    Formatea número de parte para búsqueda en parts.cat.com
    Ej: '7T1997' -> '7T-1997', '4590686' -> '459-0686'
    """
    raw = raw.strip().upper().replace(" ", "").replace("-", "")

    # Si ya tiene formato correcto, retornar
    if re.match(r'^[A-Z0-9]+-[A-Z0-9]+$', raw):
        return raw

    # Patrón letras+números: 7T1997 -> 7T-1997, 1A1135 -> 1A-1135
    m = re.match(r'^([A-Z0-9]{1,3}[A-Z])(\d{4,})$', raw)
    if m:
        return f"{m.group(1)}-{m.group(2)}"

    # Patrón solo números con 7 dígitos: 4590686 -> 459-0686
    if re.match(r'^\d{7}$', raw):
        return f"{raw[:3]}-{raw[3:]}"

    # Patrón 7 dígitos alfanumerico: 5872880 -> 587-2880
    if re.match(r'^\d{6,8}$', raw):
        mid = len(raw) // 2
        return f"{raw[:mid]}-{raw[mid:]}"

    return raw


def generate_result_excel(filepath_original: str, items_data: List[dict], output_path: str) -> str:
    """
    Genera Excel con columnas adicionales de datos CAT al final.
    Mantiene el formato original y agrega columnas: NOMBRE CAT, PRECIO CAT, RETIRO ESTIMADO, URL
    """
    wb = openpyxl.load_workbook(filepath_original)
    ws = wb.active

    # Encontrar fila de encabezados
    header_row_idx = None
    for row in ws.iter_rows():
        for cell in row:
            cv = str(cell.value).strip().upper() if cell.value else ""
            if cv and ("PARTE" in cv or "PART NUMBER" in cv or cv == "PART"):
                header_row_idx = cell.row
                break
        if header_row_idx:
            break

    if not header_row_idx:
        raise ValueError("No se encontró fila de encabezados")

    # Encontrar última columna usada en la fila de headers
    header_row = ws[header_row_idx]
    last_col = 0
    for cell in header_row:
        if cell.value is not None:
            last_col = cell.column

    new_cols_start = last_col + 1

    # Colores
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    found_fill = PatternFill("solid", fgColor="E8F5E9")
    not_found_fill = PatternFill("solid", fgColor="FFF3E0")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Nuevas columnas header
    new_headers = ["NOMBRE CAT", "PRECIO CAT (CLP)", "RETIRO ESTIMADO", "ESTADO", "URL"]
    for i, h in enumerate(new_headers):
        col = new_cols_start + i
        cell = ws.cell(row=header_row_idx, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = 22 if i < 3 else 12 if i == 3 else 35

    # Crear mapa de numero_parte -> item_data
    items_map = {item["numero_parte"]: item for item in items_data}

    # Llenar datos en filas de items
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        # Buscar columna N° PARTE en esta fila
        np_val = None
        np_col_idx = None
        for cell in row:
            if cell.value and header_row_idx:
                header_cell = ws.cell(row=header_row_idx, column=cell.column)
                hv = str(header_cell.value).strip().upper() if header_cell.value else ""
                if hv and ("PARTE" in hv or "PART NUMBER" in hv or hv == "PART"):
                    np_val = str(cell.value).strip()
                    np_col_idx = cell.column
                    break

        if not np_val or not np_val.strip():
            continue

        item_data = items_map.get(np_val)
        if not item_data:
            continue

        found = item_data.get("encontrado", 0)
        fill = found_fill if found else not_found_fill

        values = [
            item_data.get("nombre_cat", "") or "",
            item_data.get("precio_cat", "") or "",
            item_data.get("retiro_estimado", "") or "",
            "✓ Encontrado" if found else "✗ No encontrado",
            item_data.get("url_cat", "") or "",
        ]

        for i, val in enumerate(values):
            col = new_cols_start + i
            cell = ws.cell(row=row[0].row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(i == 4))
            if i == 1 and val:
                try:
                    cell.number_format = '#,##0.00'
                except:
                    pass

    wb.save(output_path)
    return output_path
